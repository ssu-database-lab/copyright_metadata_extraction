"""
Gemma vs Qwen 이미지 설명 평가 — KOGL GOLD 설명 기준.

입력: kogl_gold_scraper.py 산출 (kogl_gold.xlsx + images/). status==ok 행만 사용.
각 이미지에 대해 Gemma-4-31B(OpenRouter)와 Qwen3-VL(DashScope)로 한국어 설명을 생성하고,
KOGL GOLD 설명(저작물 설명) 기준으로 채점:
  1) LLM-as-judge (qwen-max): adequacy(내용충실 1-5) + faithfulness(환각없음/사실일치 1-5).
  2) Head-to-head: gold에 더 부합하는 설명(A/B/tie), 위치편향 방지 위해 A/B 무작위 배치.
  3) Embedding cosine (text-embedding-v3): cos(추론, gold).
자유텍스트라 정확 토큰일치(F1/BLEU)는 무효 — 위 3지표 사용.

사용:
  python -m api.module.clip_extraction.vlm.eval_gold_descriptions \
      --gold dataset/kogl_gold/kogl_gold.xlsx --images dataset/kogl_gold/images \
      --out dataset/kogl_gold/eval_gold.xlsx --n 60
"""
from __future__ import annotations
import argparse, json, os, random, re, time
from pathlib import Path
from typing import Dict, List, Optional

DESC_SYSTEM = "당신은 이미지를 한국어로 정확하게 서술하는 전문가입니다."
DESC_USER = (
    "이 이미지를 한국어로 3~5문장으로 구체적으로 설명하세요. "
    "무엇이 보이는지(사물·장소·인물·구도·분위기·맥락)를 포함하되, "
    "확실하지 않은 사실(정확한 지명·연도·고유명)은 추측하지 마세요. 설명 문장만 출력하세요."
)


def _dashscope_client():
    from dotenv import dotenv_values
    from openai import OpenAI
    env = dotenv_values(str(Path(__file__).resolve().parents[4] / ".env"))
    key = env.get("DASHSCOPE_API_KEY") or os.getenv("DASHSCOPE_API_KEY")
    base = (env.get("DASHSCOPE_BASE_URL") or os.getenv("DASHSCOPE_BASE_URL")
            or "https://dashscope-intl.aliyuncs.com/compatible-mode/v1")
    return OpenAI(api_key=key, base_url=base)


def _cosine(a: List[float], b: List[float]) -> float:
    import math
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a)); nb = math.sqrt(sum(y * y for y in b))
    return dot / (na * nb) if na and nb else 0.0


def _judge_one(client, judge_model: str, gold: str, cand: str) -> Dict:
    """(gold, 후보설명) → subject_consistency/descriptive_quality 1-5 + rationale.

    주의: KOGL GOLD는 '이미지 캡션'이 아니라 대상에 대한 배경 설명(에세이)이다.
    따라서 '에세이 재현'이 아니라 (a) 대상 부합 (b) 서술 품질로 채점한다.
    """
    prompt = (
        "GOLD는 어떤 대상(장소·문화재 등)에 대한 배경 설명 글이고, CANDIDATE는 그 대상을 찍은 "
        "사진 1장에 대해 모델이 생성한 시각적 설명입니다. 두 글은 목적이 다릅니다(GOLD=배경지식, "
        "CANDIDATE=사진묘사). CANDIDATE가 GOLD의 배경 문장을 재현하지 못하는 것은 감점 사유가 아닙니다.\n"
        "다음 두 축으로 채점하세요:\n"
        "- subject_consistency(대상 부합 1-5): CANDIDATE가 묘사한 장면·대상이 GOLD의 대상과 "
        "모순되지 않고 부합하는가. (예: GOLD가 한국 전통정원인데 CANDIDATE가 '일본식 건물'이라 하면 감점)\n"
        "- descriptive_quality(서술 품질 1-5): CANDIDATE가 사진 속 요소를 얼마나 구체적·정확·명료하게 "
        "묘사하는가(구도·사물·분위기 등). 사실 오류·환각이 있으면 감점.\n"
        "반드시 JSON 하나만 출력: {\"subject_consistency\": <1-5>, \"descriptive_quality\": <1-5>, \"rationale\": \"<한 문장>\"}\n\n"
        f"[GOLD]\n{gold}\n\n[CANDIDATE]\n{cand}"
    )
    try:
        r = client.chat.completions.create(
            model=judge_model, temperature=0, max_tokens=200,
            messages=[{"role": "system", "content": "당신은 엄격한 한국어 평가자입니다. JSON만 출력합니다."},
                      {"role": "user", "content": prompt}])
        txt = r.choices[0].message.content or ""
        m = re.search(r"\{.*\}", txt, re.S)
        d = json.loads(m.group(0)) if m else {}
        return {"subject": int(d.get("subject_consistency", 0)),
                "quality": int(d.get("descriptive_quality", 0)),
                "rationale": str(d.get("rationale", ""))[:200]}
    except Exception as e:
        return {"subject": 0, "quality": 0, "rationale": f"judge_err:{type(e).__name__}"}


def _judge_h2h(client, judge_model: str, gold: str, a: str, b: str) -> str:
    """gold에 더 부합하는 설명(A/B/tie)."""
    prompt = (
        "GOLD는 대상에 대한 배경 설명이고 A/B는 그 대상 사진에 대한 시각적 설명입니다. "
        "대상에 더 부합하면서(모순·환각 없음) 더 구체적·정확하게 사진을 묘사한 쪽을 고르세요. "
        "반드시 JSON 하나만: {\"winner\": \"A\"|\"B\"|\"tie\"}\n\n"
        f"[GOLD]\n{gold}\n\n[A]\n{a}\n\n[B]\n{b}"
    )
    try:
        r = client.chat.completions.create(
            model=judge_model, temperature=0, max_tokens=20,
            messages=[{"role": "user", "content": prompt}])
        m = re.search(r'"winner"\s*:\s*"(A|B|tie)"', r.choices[0].message.content or "")
        return m.group(1) if m else "tie"
    except Exception:
        return "tie"


def _embed(client, model: str, texts: List[str]) -> List[List[float]]:
    r = client.embeddings.create(model=model, input=texts)
    return [d.embedding for d in r.data]


def main() -> int:
    ap = argparse.ArgumentParser(description="Gemma vs Qwen 설명 평가 (KOGL GOLD 기준)")
    ap.add_argument("--gold", default="dataset/kogl_gold/kogl_gold.xlsx")
    ap.add_argument("--images", default="dataset/kogl_gold/images")
    ap.add_argument("--out", default="dataset/kogl_gold/eval_gold.xlsx")
    ap.add_argument("--n", type=int, default=0, help="평가 이미지 수 상한(0=전체)")
    ap.add_argument("--judge", default="qwen-max")
    ap.add_argument("--embed", default="text-embedding-v3")
    ap.add_argument("--max-tokens", type=int, default=400)
    ap.add_argument("--seed", type=int, default=13)
    args = ap.parse_args()

    import openpyxl
    from api.module.clip_extraction.vlm.extractor import make_openrouter_backend, make_qwen_backend
    random.seed(args.seed)

    wb = openpyxl.load_workbook(args.gold); ws = wb.active
    rows = list(ws.iter_rows(values_only=True)); hdr = rows[0]
    ci = {h: i for i, h in enumerate(hdr)}
    items = [r for r in rows[1:] if r[ci["status"]] == "ok"]
    if args.n:
        items = items[: args.n]
    img_dir = Path(args.images)
    print(f"평가 대상 {len(items)}건 | judge={args.judge} embed={args.embed}")

    gemma = make_openrouter_backend()
    qwen = make_qwen_backend()
    ds = _dashscope_client()

    def describe(backend, img: Path) -> str:
        try:
            res = backend.extract(img, DESC_SYSTEM, DESC_USER,
                                  max_tokens=args.max_tokens, temperature=0.0)
            return (res.raw_text or "").strip() if res and res.ok else f"[FAIL:{getattr(res,'error','')}]"
        except Exception as e:
            return f"[ERR:{type(e).__name__}:{e}]"

    results = []
    for i, r in enumerate(items, 1):
        idx = r[ci["recommendIdx"]]; gold = str(r[ci["gold_desc"]])
        img = img_dir / str(r[ci["image_file"]])
        if not img.exists():
            print(f"  ✗ [{i}] {idx} 이미지 없음"); continue
        g_desc = describe(gemma, img)
        q_desc = describe(qwen, img)
        gj = _judge_one(ds, args.judge, gold, g_desc)
        qj = _judge_one(ds, args.judge, gold, q_desc)
        # head-to-head, 위치편향 방지
        swap = random.random() < 0.5
        a, b = (g_desc, q_desc) if not swap else (q_desc, g_desc)
        w = _judge_h2h(ds, args.judge, gold, a, b)
        winner = {"tie": "tie"}.get(w, ("gemma" if (w == "A") != swap else "qwen"))
        # cosine
        try:
            embs = _embed(ds, args.embed, [gold, g_desc, q_desc])
            g_cos = _cosine(embs[0], embs[1]); q_cos = _cosine(embs[0], embs[2])
        except Exception:
            g_cos = q_cos = 0.0
        rec = {"recommendIdx": idx, "title": r[ci["title"]], "gold_len": len(gold),
               "gold_desc": gold, "gemma_desc": g_desc, "qwen_desc": q_desc,
               "gemma_subj": gj["subject"], "gemma_qual": gj["quality"], "gemma_cos": round(g_cos, 4),
               "qwen_subj": qj["subject"], "qwen_qual": qj["quality"], "qwen_cos": round(q_cos, 4),
               "h2h_winner": winner, "gemma_rat": gj["rationale"], "qwen_rat": qj["rationale"]}
        results.append(rec)
        print(f"  ✅ [{i}/{len(items)}] {idx} | Gemma s{gj['subject']}/q{gj['quality']} cos{g_cos:.2f} "
              f"| Qwen s{qj['subject']}/q{qj['quality']} cos{q_cos:.2f} | win={winner}")

    # write + aggregate
    owb = openpyxl.Workbook(); ows = owb.active; ows.title = "per_image"
    cols = ["recommendIdx", "title", "gold_len", "gemma_subj", "gemma_qual", "gemma_cos",
            "qwen_subj", "qwen_qual", "qwen_cos", "h2h_winner",
            "gold_desc", "gemma_desc", "qwen_desc", "gemma_rat", "qwen_rat"]
    ows.append(cols)
    for r in results:
        ows.append([r.get(c, "") for c in cols])

    def mean(key):
        vals = [r[key] for r in results if isinstance(r[key], (int, float)) and r[key] > 0]
        return round(sum(vals) / len(vals), 3) if vals else 0.0
    agg = owb.create_sheet("aggregate")
    n = len(results)
    wins = {"gemma": sum(1 for r in results if r["h2h_winner"] == "gemma"),
            "qwen": sum(1 for r in results if r["h2h_winner"] == "qwen"),
            "tie": sum(1 for r in results if r["h2h_winner"] == "tie")}
    g_err = sum(1 for r in results if isinstance(r["gemma_subj"], int) and 0 < r["gemma_subj"] < 3)
    q_err = sum(1 for r in results if isinstance(r["qwen_subj"], int) and 0 < r["qwen_subj"] < 3)
    agg.append(["metric", "Gemma-4-31B", "Qwen3-VL"])
    agg.append(["mean subject_consistency (1-5)", mean("gemma_subj"), mean("qwen_subj")])
    agg.append(["mean descriptive_quality (1-5)", mean("gemma_qual"), mean("qwen_qual")])
    agg.append(["mean cosine vs gold", mean("gemma_cos"), mean("qwen_cos")])
    agg.append(["subject-error (subj<3) count", g_err, q_err])
    agg.append(["head-to-head wins", wins["gemma"], wins["qwen"]])
    agg.append(["h2h ties", wins["tie"], ""])
    agg.append(["n images", n, ""])
    owb.save(args.out)

    print("\n" + "=" * 64)
    print(f"  집계 (n={n})                 Gemma-4-31B    Qwen3-VL")
    print(f"  mean subject_consistency(1-5)  {mean('gemma_subj'):>6}       {mean('qwen_subj'):>6}")
    print(f"  mean descriptive_quality(1-5)  {mean('gemma_qual'):>6}       {mean('qwen_qual'):>6}")
    print(f"  mean cosine vs gold            {mean('gemma_cos'):>6}       {mean('qwen_cos'):>6}")
    print(f"  subject-error (subj<3)         {g_err:>6}       {q_err:>6}")
    print(f"  head-to-head wins              {wins['gemma']:>6}       {wins['qwen']:>6}   (tie {wins['tie']})")
    print("=" * 64)
    print(f"  결과: {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""
Score a vlm_compare report (Gemma vs Qwen) into a P1 baseline:
  - work_type distribution per model + Gemma-vs-Qwen agreement rate
  - keyword overlap vs manifest weak labels (주제어 ∪ 해시태그), set-based P/R/F1
  - parse-success + latency summary

Usage:
  python -m api.module.clip_extraction.vlm.score_baseline <vlm_compare_*.json> [--manifest dataset/manifest.xlsx]
Writes a markdown summary next to the input report (…_baseline.md) and prints it.
"""
from __future__ import annotations
import argparse, json, re
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parents[4]


def _toks(x):
    """Normalize a keyword field (string or list) into a set of tokens."""
    if not x:
        return set()
    if isinstance(x, str):
        parts = re.split(r"[,\s/#·]+", x)
    elif isinstance(x, list):
        parts = []
        for it in x:
            parts += re.split(r"[,\s/#·]+", str(it))
    else:
        parts = [str(x)]
    return {p.strip().lower() for p in parts if p.strip()}


def _load_manifest_labels(manifest_path: Path):
    """idx -> {keywords:set, 공공누리유형, 정보유형} from manifest weak labels."""
    import openpyxl
    ws = openpyxl.load_workbook(manifest_path).active
    hdr = [c.value for c in ws[1]]; ci = {h: i for i, h in enumerate(hdr)}
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        idx = str(r[ci["원문인덱스"]])
        kw = _toks(r[ci.get("주제어")]) | _toks(r[ci.get("해시태그")])
        out[idx] = {"keywords": kw,
                    "유형": r[ci.get("공공누리 유형")] if "공공누리 유형" in ci else None}
    return out


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("report")
    ap.add_argument("--manifest", default=str(ROOT / "dataset" / "manifest.xlsx"))
    args = ap.parse_args(argv)
    rep = json.loads(Path(args.report).read_text(encoding="utf-8"))
    labels = _load_manifest_labels(Path(args.manifest)) if Path(args.manifest).exists() else {}
    models = rep["models"]
    by_image = rep["by_image"]

    wt = {m: Counter() for m in models}        # work_type distribution
    kw_f1 = {m: [] for m in models}            # keyword set-F1 vs weak labels
    agree = 0; comparable = 0
    gemma_lbl = next((m for m in models if m.startswith("Gemma")), None)
    qwen_lbl = next((m for m in models if m.startswith("Qwen")), None)

    for img, recs in by_image.items():
        idx = Path(img).stem
        gold = labels.get(idx, {}).get("keywords", set())
        per = {}
        for m in models:
            rec = recs.get(m) or {}
            p = rec.get("parsed") or {}
            w = p.get("work_type"); per[m] = w
            if w: wt[m][w] += 1
            if gold:
                pred = _toks(p.get("keywords") or p.get("main_subjects"))
                if pred:
                    inter = len(pred & gold)
                    prec = inter / len(pred); rec_ = inter / len(gold)
                    f1 = (2*prec*rec_/(prec+rec_)) if (prec+rec_) else 0.0
                    kw_f1[m].append(f1)
        if gemma_lbl and qwen_lbl and per.get(gemma_lbl) and per.get(qwen_lbl):
            comparable += 1
            if per[gemma_lbl] == per[qwen_lbl]: agree += 1

    L = [f"# VLM P1 Baseline — {rep['timestamp']}",
         f"\n- images: {rep['n_images']} | models: {', '.join(models)} | manifest labels: {len(labels)} works",
         f"- weak-label keyword coverage: {sum(1 for i in by_image if labels.get(Path(i).stem,{}).get('keywords'))}/{len(by_image)} images have 주제어/해시태그\n",
         "## Parse + latency", "| Model | OK | parsed | avg latency | avg tokens |", "|---|---|---|---|---|"]
    for m, agg in rep["summary"].items():
        L.append(f"| {m} | {agg['ok']}/{rep['n_images']} | {agg['parsed']}/{rep['n_images']} | {agg['avg_latency']}s | {agg['avg_completion_tokens']} |")
    L.append("\n## work_type distribution")
    for m in models:
        L.append(f"- **{m}**: {dict(wt[m])}")
    if gemma_lbl and qwen_lbl:
        L.append(f"\n## Gemma-vs-Qwen work_type agreement: **{agree}/{comparable} = {(100*agree/comparable if comparable else 0):.1f}%**")
    L.append("\n## Keyword set-F1 vs manifest 주제어/해시태그 (weak labels)")
    for m in models:
        sc = kw_f1[m]
        L.append(f"- **{m}**: mean F1 {sum(sc)/len(sc):.3f} over {len(sc)} labeled images" if sc else f"- **{m}**: no labeled images to score")
    out = Path(args.report).with_name(Path(args.report).stem + "_baseline.md")
    out.write_text("\n".join(L), encoding="utf-8")
    print("\n".join(L)); print(f"\n[written] {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

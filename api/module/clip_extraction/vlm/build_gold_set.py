"""
Build a small GOLD evaluation set for VLM attribute extraction (design §8).

Produces a human-labeling sheet: each row = one image with Gemma + Qwen
predictions (work_type / keywords / description) and EMPTY GOLD_* columns for a
human to fill. Reuses an existing vlm_compare report (no re-run for those
images) and optionally adds N more fresh images. Copies the gold images into a
local folder so the labeler can view them.

Usage:
  python -m api.module.clip_extraction.vlm.build_gold_set \
     --report api/module/clip_extraction/reports/vlm_compare_YYYY..json \
     --add 50 --gemma-url http://100.105.203.69:8001/v1
Output: dataset/gold_set/{images/, gold_to_label.xlsx}
Then a human fills GOLD_work_type / GOLD_desc_grade(1=wrong,2=partial,3=good) /
GOLD_keyword_precision(0-1) / GOLD_notes, and score_gold.py computes the KPIs.
"""
from __future__ import annotations
import argparse, json, shutil, sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(ROOT / "api"))
IMG_DIR = Path("/mnt/e/kogl_originals/이미지")


def _vlm_fields(rec: dict) -> dict:
    p = (rec or {}).get("parsed") or {}
    kws = p.get("keywords") or p.get("main_subjects") or []
    return {"work_type": p.get("work_type"),
            "keywords": ", ".join(kws) if isinstance(kws, list) else str(kws or ""),
            "description": p.get("description")}


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--report", default="", help="existing vlm_compare_*.json to reuse")
    ap.add_argument("--images", default=str(IMG_DIR))
    ap.add_argument("--add", type=int, default=50, help="extra fresh images to run + add")
    ap.add_argument("--gemma-url", default="http://100.105.203.69:8001/v1")
    ap.add_argument("--out", default=str(ROOT / "dataset" / "gold_set"))
    args = ap.parse_args(argv)

    import openpyxl
    from api.module.clip_extraction.vlm.compare import build_clients, discover_images
    from api.module.clip_extraction.vlm.prompts import get_prompts
    sys_p, usr_p = get_prompts("ko")

    # manifest hints (제목 / 정보유형 / 주제어) by idx
    man = {}
    mpath = ROOT / "dataset" / "manifest.xlsx"
    if mpath.exists():
        ws = openpyxl.load_workbook(mpath).active
        h = [c.value for c in ws[1]]; ci = {k: i for i, k in enumerate(h)}
        for r in ws.iter_rows(min_row=2, values_only=True):
            man[str(r[ci["원문인덱스"]])] = {
                "title": r[ci["제목"]],
                "정보유형": r[ci.get("정보유형")] if "정보유형" in ci else None,
                "주제어": r[ci.get("주제어")] if "주제어" in ci else None,
            }

    rows = {}  # img_name -> {gemma:{}, qwen:{}}
    used = set()
    # 1) reuse report predictions
    if args.report and Path(args.report).exists():
        rep = json.loads(Path(args.report).read_text(encoding="utf-8"))
        gem = next((m for m in rep["models"] if m.startswith("Gemma")), None)
        qw = next((m for m in rep["models"] if m.startswith("Qwen")), None)
        for img, recs in rep["by_image"].items():
            rows[img] = {"gemma": _vlm_fields(recs.get(gem)), "qwen": _vlm_fields(recs.get(qw))}
            used.add(img)
        print(f"reused {len(rows)} images from report")

    # 2) add N fresh images (evenly spaced, not already used)
    all_imgs = discover_images(Path(args.images))
    pool = [p for p in all_imgs if p.name not in used]
    add = []
    if args.add and pool:
        step = max(1, len(pool) // args.add)
        add = [pool[i] for i in range(0, len(pool), step)][: args.add]
    if add:
        clients = build_clients(["gemma", "qwen"], args.gemma_url)
        live = [c for c in clients if c.ping()[0]]
        print(f"running {len(live)} model(s) on {len(add)} new images...")
        for i, ip in enumerate(add, 1):
            rows[ip.name] = {}
            for c in live:
                res = c.extract(ip, sys_p, usr_p, max_tokens=2048, temperature=0.0)
                key = "gemma" if c.model_label.startswith("Gemma") else "qwen"
                rows[ip.name][key] = _vlm_fields(res.__dict__)
            if i % 10 == 0:
                print(f"  {i}/{len(add)}")

    # 3) copy images + write sheet
    out = Path(args.out); (out / "images").mkdir(parents=True, exist_ok=True)
    name_to_path = {p.name: p for p in all_imgs}
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "gold"
    cols = ["원문인덱스", "제목", "image_file", "manifest_정보유형", "manifest_주제어",
            "gemma_work_type", "qwen_work_type", "agree",
            "gemma_keywords", "gemma_description", "qwen_description",
            "GOLD_work_type", "GOLD_desc_grade(1-3)", "GOLD_keyword_precision(0-1)", "GOLD_notes"]
    ws.append(cols)
    n = 0
    for img in sorted(rows):
        idx = Path(img).stem; src = name_to_path.get(img)
        if src and src.exists():
            shutil.copy(src, out / "images" / img)
        g = rows[img].get("gemma", {}); q = rows[img].get("qwen", {})
        mh = man.get(idx, {})
        agree = "Y" if (g.get("work_type") and g.get("work_type") == q.get("work_type")) else "N"
        ws.append([idx, mh.get("title"), img, mh.get("정보유형"), mh.get("주제어"),
                   g.get("work_type"), q.get("work_type"), agree,
                   g.get("keywords"), g.get("description"), q.get("description"),
                   "", "", "", ""])
        n += 1
    wb.save(out / "gold_to_label.xlsx")
    print(f"\nGOLD set: {n} images → {out}/gold_to_label.xlsx (+ images/)")
    print("Human fills GOLD_work_type / GOLD_desc_grade(1=wrong,2=partial,3=good) / "
          "GOLD_keyword_precision(0-1) / GOLD_notes, then run score_gold.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

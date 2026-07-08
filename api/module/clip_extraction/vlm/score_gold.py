"""
Score a human-filled gold sheet (from build_gold_set.py) into VLM KPIs:
  - work_type accuracy: Gemma vs GOLD, Qwen vs GOLD (overall + per-type)
  - description adequacy: distribution of GOLD_desc_grade (1=wrong,2=partial,3=good) + mean
  - keyword precision: mean GOLD_keyword_precision (Gemma keywords judged 0-1)

Only rows where GOLD_work_type is filled are scored.

Usage: python -m api.module.clip_extraction.vlm.score_gold dataset/gold_set/gold_to_label.xlsx
"""
from __future__ import annotations
import argparse
from collections import Counter, defaultdict
from pathlib import Path


def main(argv=None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("sheet")
    args = ap.parse_args(argv)
    import openpyxl
    ws = openpyxl.load_workbook(args.sheet).active
    hdr = [c.value for c in ws[1]]; ci = {h: i for i, h in enumerate(hdr)}

    def g(r, k):
        return r[ci[k]] if k in ci and ci[k] < len(r) else None

    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    labeled = [r for r in rows if str(g(r, "GOLD_work_type") or "").strip()]
    print(f"# VLM Gold-set scores — {Path(args.sheet).name}")
    print(f"- rows: {len(rows)} | human-labeled: {len(labeled)}")
    if not labeled:
        print("\nNo GOLD_work_type filled yet — fill the sheet then re-run.")
        return 0

    # work_type accuracy
    for model_col, name in [("gemma_work_type", "Gemma"), ("qwen_work_type", "Qwen")]:
        if model_col not in ci:
            continue
        correct = 0; per = defaultdict(lambda: [0, 0])  # gold_type -> [correct, total]
        for r in labeled:
            gold = str(g(r, "GOLD_work_type")).strip()
            pred = str(g(r, model_col) or "").strip()
            per[gold][1] += 1
            if pred == gold:
                correct += 1; per[gold][0] += 1
        acc = 100 * correct / len(labeled)
        print(f"\n## {name} work_type accuracy: {correct}/{len(labeled)} = {acc:.1f}%")
        for t, (c, n) in sorted(per.items(), key=lambda x: -x[1][1]):
            print(f"   {t}: {c}/{n} ({100*c/n:.0f}%)")

    # description adequacy
    grades = [int(g(r, "GOLD_desc_grade(1-3)")) for r in labeled
              if str(g(r, "GOLD_desc_grade(1-3)") or "").strip().isdigit()]
    if grades:
        dist = Counter(grades)
        print(f"\n## Description adequacy (Gemma): mean {sum(grades)/len(grades):.2f}/3 "
              f"over {len(grades)} | dist 3(good)={dist.get(3,0)} 2(partial)={dist.get(2,0)} 1(wrong)={dist.get(1,0)}")

    # keyword precision
    kp = []
    for r in labeled:
        v = g(r, "GOLD_keyword_precision(0-1)")
        try:
            kp.append(float(v))
        except (TypeError, ValueError):
            pass
    if kp:
        print(f"\n## Keyword precision (Gemma): mean {sum(kp)/len(kp):.3f} over {len(kp)} labeled")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

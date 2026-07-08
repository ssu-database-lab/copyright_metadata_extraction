"""
KOGL 원본 파일 다운로더 (인증 필요) — Playwright로 SSO 로그인 1회 후, 작품별로 다운로드 팝업을
열어 활용용도(Z=기타)+약관동의를 채우고 제출하여 원본 파일을 캡처한다.

검증된 흐름(2026-06-19): /personalpage/personalMain.do → SSO SPA '아이디 로그인' → #id/#pw →
#btn-login → KOGL 인증됨 → 상세페이지 '이 저작물만 다운로드' 팝업 → recomDownloadFile → 원본.

- 인증: KOGL_ID / KOGL_PW 환경변수 (또는 .env). 평문 비밀번호는 코드/로그에 남기지 말 것.
- 영상(vid)은 50MB 초과 시 건너뜀(기본). 가능하면 다운로드 전 크기 표시를 읽어 스킵,
  못 읽으면 받은 뒤 크기 확인 후 삭제.
- 분류 → division: 이미지→img, 어문→doc, 영상→vid.
- 재개 가능(기존 파일 스킵), originals_index.xlsx에 상태 기록.

사용:
    KOGL_ID=... KOGL_PW=... python -m api.module.dataset_builder.download_originals --dry-run 10
    KOGL_ID=... KOGL_PW=... python -m api.module.dataset_builder.download_originals --types 이미지,어문,영상
"""
from __future__ import annotations
import argparse, os, re, time
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
DEFAULT_MANIFEST = ROOT / "dataset" / "manifest.xlsx"
DEFAULT_OUT = ROOT / "dataset" / "originals"
ENTRY = "https://www.kogl.or.kr/personalpage/personalMain.do"
DETAIL = "https://www.kogl.or.kr/recommend/recommendDivView.do?recommendIdx={idx}&division={div}"
# division 코드: 상세페이지 URL ?division= 값. 영상은 'video'(‘vid’ 아님 — 잘못되면 다운로드 컨트롤 미표시).
DIVISION = {"이미지": "img", "어문": "doc", "영상": "video", "오디오": "sound", "3D": "3D", "글꼴": "font"}
# 주의: purpose 가 너무 길면(≈20자 초과) 서버가 "파일을 찾을수 없습니다" 를 반환한다.
# maxlength 속성은 없지만 서버가 길이를 제한함(2026-06-19 확인: 19자 OK, 28자 실패). 짧게 유지.
PURPOSE = "연구개발 학습데이터셋 구축"
# 다운로드 팝업 여는 컨트롤(분류 무관): 이미지는 <a>'이 저작물만 다운로드', 어문은 <button>'다운로드'
# 둘 다 onclick="recomOpenPop('popDownload_File', this)" 를 가진다. 영상은 이 컨트롤이 없음(원본 다운 불가).
OPENER = "[onclick*=\"recomOpenPop('popDownload_File'\"]"

FILL_JS = """(purpose) => {
  // recomDownloadFile 가 실제로 검증하는 값만 채운다: useType(=Z면 #purpose) + #agree.
  // change 이벤트만 발생(합성 click 은 #agree onclick 토글을 다시 호출해 체크를 풀어버림).
  // #applicationItemA~D 등 부가 체크박스는 파일 선택 컨트롤로 보이며, 강제 체크 시
  // 존재하지 않는 파일 슬롯을 가리켜 서버가 "파일을 찾을수 없습니다" 를 반환한다(2026-06-19 확인).
  const fire=(e)=>e.dispatchEvent(new Event('change',{bubbles:true}));
  let z=document.querySelector("input[name=useType][value='Z']");
  if(z){z.checked=true;fire(z);const pp=document.querySelector('#purpose');if(pp){pp.value=purpose||'연구';fire(pp);}}
  else{let r=document.querySelector("input[name=useType]");if(r){r.checked=true;fire(r);let s=document.querySelector("input[name=useTypeSub]");if(s){s.checked=true;fire(s);}}}
  ['#agree','#agreeAll'].forEach(sel=>document.querySelectorAll(sel).forEach(e=>{e.checked=true;fire(e);}));
}"""


def login(pg, uid: str, pw: str) -> None:
    pg.goto(ENTRY, wait_until="domcontentloaded", timeout=45000)
    pg.wait_for_timeout(4500)
    pg.get_by_text("아이디 로그인", exact=False).first.click(timeout=15000)
    pg.wait_for_timeout(2500)
    pg.fill("#id", uid)
    pg.fill("#pw", pw)
    pg.locator("#btn-login").click()
    for _ in range(3):
        try:
            pg.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass
    pg.wait_for_timeout(1500)


def _size_from_popup(pg) -> int | None:
    """팝업/페이지에서 파일 크기(…MB/KB) 베스트에포트 추출. 못 찾으면 None."""
    try:
        txt = pg.evaluate("""() => {
            const els=[...document.querySelectorAll('#popDownload_File *,.dataFileIdx,[class*=file]')];
            return els.map(e=>e.innerText||'').filter(t=>/\\d[\\d.,]*\\s*(MB|KB|GB|바이트|byte)/i.test(t)).slice(0,5).join(' | ');
        }""")
        if not txt:
            return None
        m = re.search(r"([\d.,]+)\s*(GB|MB|KB)", txt, re.I)
        if not m:
            return None
        v = float(m.group(1).replace(",", "")); unit = m.group(2).upper()
        return int(v * {"GB": 1024**3, "MB": 1024**2, "KB": 1024}[unit])
    except Exception:
        return None


def _attempt(pg, idx: str, division: str, out_dir: Path, video_cap: int) -> dict:
    """한 번의 다운로드 시도. 단말 상태(ok/skipped_video_large/no_in_site_download/
    not_authenticated)면 그대로 반환, 일시적 실패는 status='retry' + error 로 반환."""
    # 영상 상세는 플레이어가 계속 폴링해 networkidle 에 도달하지 못함 → domcontentloaded 사용
    pg.goto(DETAIL.format(idx=idx, div=division), wait_until="domcontentloaded", timeout=45000)
    pg.wait_for_timeout(1500)
    if not (("로그아웃" in pg.content()) or ("마이페이지" in pg.content())):
        return {"status": "not_authenticated"}
    opener = pg.locator(OPENER)
    if opener.count() == 0:
        # 원본 다운로드 컨트롤 없음(주로 영상) → 메타데이터만 보유
        return {"status": "no_in_site_download"}
    opener.first.click(timeout=10000)
    pg.wait_for_timeout(1800)
    pg.evaluate(FILL_JS, PURPOSE)
    pg.wait_for_timeout(400)
    alert = {}
    clen = {}  # 다운로드 응답의 Content-Length (영상 대용량 사전 차단용)
    def on_resp(r):
        try:
            cd = r.headers.get("content-disposition", "")
            if ("attachment" in cd) or ("recommFileDown" in r.url):
                cl = r.headers.get("content-length")
                if cl and cl.isdigit():
                    clen["n"] = int(cl)
        except Exception:
            pass
    dlg = lambda d: (alert.setdefault("m", d.message), d.accept())
    pg.on("dialog", dlg)
    pg.on("response", on_resp)
    try:
        with pg.expect_download(timeout=30000) as di:
            # '이 저작물만 다운로드' 제출 버튼만 정확히 타겟(recomDownloadFile). '전체 다운로드'
            # (recomDownloadAll) 버튼이 함께 있는 페이지에서 .last 가 숨겨진 All 버튼을 잡아
            # 8s 타임아웃 나던 문제 수정(2026-06-19). 짧게(8s)로 비다운로드 작품 낭비도 방지.
            sub = pg.locator("button[onclick*='recomDownloadFile']")
            (sub.first if sub.count() else
             pg.locator("button[type=submit]:has-text('다운로드'), button:has-text('다운로드')").last).click(timeout=8000)
        d = di.value
        pg.wait_for_timeout(300)  # 응답 헤더(Content-Length) 도착 대기
        # 영상: 전체 바이트 수신 전에 Content-Length 로 사전 차단. video_cap=0 은 무제한.
        if video_cap and division == "video" and clen.get("n", 0) > video_cap:
            try: d.cancel()
            except Exception: pass
            return {"status": "skipped_video_large", "size": clen["n"]}
        ext = os.path.splitext(d.suggested_filename or "")[1] or ".bin"
        path = out_dir / f"{idx}{ext}"
        d.save_as(str(path))
        size = path.stat().st_size
        # 사후 안전망: Content-Length 를 못 읽은 경우. video_cap=0 은 무제한.
        if video_cap and division == "video" and size > video_cap:
            path.unlink(missing_ok=True)
            return {"status": "skipped_video_large", "size": size}
        return {"status": "ok", "file": path.name, "size": size}
    except Exception as e:
        # 서버가 "파일을 찾을수 없습니다" alert 를 띄우거나(간헐적) 다운로드 미발생 → 재시도 대상
        return {"status": "retry", "error": (alert.get("m") or str(e))[:160]}
    finally:
        pg.remove_listener("dialog", dlg)
        pg.remove_listener("response", on_resp)


def download_one(pg, idx: str, division: str, out_dir: Path, video_cap: int, retries: int = 3) -> dict:
    res = {"idx": idx, "division": division, "status": "", "file": "", "size": 0}
    # resume: skip if an original already exists for this idx
    existing = list(out_dir.glob(f"{idx}.*"))
    if existing and existing[0].stat().st_size > 0:
        res.update(status="skipped_existing", file=existing[0].name, size=existing[0].stat().st_size)
        return res
    last_err, last_status = "", ""
    for attempt in range(1, retries + 1):
        try:
            r = _attempt(pg, idx, division, out_dir, video_cap)
        except Exception as e:
            r = {"status": "retry", "error": str(e)[:160]}
        st = r["status"]
        if st in ("ok", "skipped_video_large", "no_in_site_download"):
            res.update(r); return res
        last_status, last_err = st, r.get("error", st)
        if attempt < retries:
            time.sleep(1.5)  # KOGL 간헐 오류 흡수
    # 재시도 소진
    res.update(status=("not_authenticated" if last_status == "not_authenticated" else "failed"),
               error=last_err)
    return res


def main() -> int:
    ap = argparse.ArgumentParser(description="KOGL 원본 파일 다운로더 (Playwright 인증)")
    ap.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--types", default="이미지,어문",
                    help="분류 (쉼표구분). 영상 기본 제외: 원본 대부분 50MB↑·사전 크기확인 불가, "
                         "메타데이터만 manifest.xlsx 보유(2026-06-19 결정)")
    ap.add_argument("--dry-run", type=int, default=0, help="유형별 N건만 (stratified)")
    ap.add_argument("--video-max-mb", type=int, default=50)
    ap.add_argument("--doc-cap-gb", type=float, default=2.0,
                    help="어문(PDF) 누적 용량 상한(GB). 초과 시 이후 어문은 다운로드하지 않고 스킵. 0=무제한")
    ap.add_argument("--limit", type=int, default=0, help="전체 상한")
    args = ap.parse_args()

    uid, pw = os.environ.get("KOGL_ID"), os.environ.get("KOGL_PW")
    if not uid or not pw:
        print("KOGL_ID / KOGL_PW 환경변수가 필요합니다 (.env 권장).")
        return 2
    video_cap = args.video_max_mb * 1024 * 1024
    types = [t.strip() for t in args.types.split(",") if t.strip()]

    import openpyxl
    ws = openpyxl.load_workbook(args.manifest).active
    hdr = [c.value for c in ws[1]]; ci = {h: i for i, h in enumerate(hdr)}
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[ci["분류"]] in types]
    # select
    if args.dry_run:
        sel = []
        for t in types:
            sel += [r for r in rows if r[ci["분류"]] == t][: args.dry_run]
        rows = sel
    elif args.limit:
        rows = rows[: args.limit]

    out = Path(args.out)
    doc_cap = int(args.doc_cap_gb * 1024**3) if args.doc_cap_gb > 0 else 0
    # 재개 대응: 기존 어문 파일 누적 용량으로 초기화
    doc_dir = out / "어문"
    doc_bytes = sum(f.stat().st_size for f in doc_dir.glob("*.*")) if doc_dir.exists() else 0
    from collections import Counter
    plan = Counter(r[ci["분류"]] for r in rows)
    cap_msg = f"어문 누적 {args.doc_cap_gb}GB 초과 스킵(현재 {doc_bytes/1024**3:.2f}GB)" if doc_cap else "어문 무제한"
    print(f"대상: {len(rows)}건 {dict(plan)} | 영상 {args.video_max_mb}MB 초과 스킵 | {cap_msg} | 출력 {out}")

    from playwright.sync_api import sync_playwright
    results = []
    with sync_playwright() as p:
        b = p.chromium.launch(headless=True, args=["--no-sandbox"])
        ctx = b.new_context(accept_downloads=True, ignore_https_errors=True,
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148 Safari/537.36")
        pg = ctx.new_page()
        print("로그인 중...")
        login(pg, uid, pw)
        # 로그인 직후 SSO 리다이렉트가 끝나기를 기다린 뒤 검증(다른 네비게이션과의 경합 방지)
        pg.wait_for_timeout(2500)
        ok = False
        for _ in range(3):
            try:
                pg.goto("https://www.kogl.or.kr/", wait_until="domcontentloaded", timeout=30000)
                ok = ("로그아웃" in pg.content()) or ("마이페이지" in pg.content())
                break
            except Exception:
                pg.wait_for_timeout(2000)
        print("로그인 OK" if ok else "⚠ 로그인 확인 실패 — 자격증명/SSO 흐름 점검 필요")
        for i, r in enumerate(rows, 1):
            cls = r[ci["분류"]]
            idx = str(r[ci["원문인덱스"]]); div = DIVISION.get(cls, "img")
            sub = out / cls
            sub.mkdir(parents=True, exist_ok=True)
            # 어문 누적 용량 상한 도달 시 다운로드 없이 스킵(기존 파일은 skipped_existing 로 정확히 기록)
            if cls == "어문" and doc_cap and doc_bytes >= doc_cap:
                ex = list(sub.glob(f"{idx}.*"))
                if ex and ex[0].stat().st_size > 0:
                    res = {"idx": idx, "division": div, "status": "skipped_existing",
                           "file": ex[0].name, "size": ex[0].stat().st_size}
                else:
                    res = {"idx": idx, "division": div, "status": "skipped_doc_cap", "file": "", "size": 0}
            else:
                res = download_one(pg, idx, div, sub, video_cap)
                # 세션 만료 대응: 인증 실패 시 1회 재로그인 후 재시도(장시간 실행 안정성)
                if res["status"] == "not_authenticated":
                    print("  🔁 세션 만료 감지 — 재로그인 후 재시도")
                    try:
                        login(pg, uid, pw); pg.wait_for_timeout(1500)
                    except Exception:
                        pass
                    res = download_one(pg, idx, div, sub, video_cap)
                # 어문 성공 시 누적 용량 갱신
                if cls == "어문" and res["status"] in ("ok", "skipped_existing"):
                    doc_bytes += res.get("size", 0)
            res["분류"] = cls; res["제목"] = r[ci["제목"]]
            results.append(res)
            mark = {"ok": "✅", "skipped_existing": "·", "skipped_video_large": "⏭",
                    "skipped_doc_cap": "📦", "no_in_site_download": "🚫", "failed": "✗",
                    "not_authenticated": "🔒"}.get(res["status"], "?")
            err = f" :: {res['error']}" if res.get("error") else ""
            print(f"  {mark} [{i}/{len(rows)}] {res['분류']} idx={idx} {res['status']} {res.get('size',0)//1024}KB {res.get('file','')}{err}")
            time.sleep(0.8)  # politeness
        b.close()

    # index + summary
    iwb = openpyxl.Workbook(); iws = iwb.active; iws.title = "originals_index"
    cols = ["원문인덱스", "분류", "제목", "status", "file", "size", "error"]
    iws.append(cols)
    for r in results:
        iws.append([r.get("idx"), r.get("분류"), r.get("제목"), r.get("status"), r.get("file"), r.get("size"), r.get("error", "")])
    out.mkdir(parents=True, exist_ok=True)
    iwb.save(out / "originals_index.xlsx")
    from collections import Counter
    summ = Counter(r["status"] for r in results)
    print(f"\n=== 요약 ===\n  {dict(summ)}")
    print(f"  원본: {out}/<분류>/*  | 인덱스: {out}/originals_index.xlsx")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

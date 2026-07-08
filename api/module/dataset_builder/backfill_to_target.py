"""
백필러 — 분류별 '다운로드 가능한' 원본을 목표치(기본 500)까지 채운다.

흐름:
  1) RETRY  : 매니페스트 작품 중 디스크에 파일 없는 것(no_in_site/실패)을 1회 재시도. 일시적
              실패는 여기서 회수된다(전이성).
  2) BACKFILL: 그래도 부족하면 144k Excel(붙임1)에서 미사용 후보를 뽑아 다운로드 시도. 성공분만
              채택, 실패/비다운로드는 tried 로 기록하고 다음 후보로(영상은 ~17~25%가 비다운로드).
  3) RECONCILE: 새 manifest = 분류별 '다운로드된' 작품 = (원본 다운로드분) + (백필 성공분).
              비다운로드 작품은 매니페스트에서 제외. 백필 작품 썸네일 받고, manifest.xlsx 재작성.
              (계약서/인덱스 재생성은 별도 단계: generate_contracts + 인덱스 빌드.)

단일 KOGL 세션을 사용하므로 메인 다운로드가 끝난 뒤에 실행할 것. 재개 가능(state json).

사용:
  KOGL_ID=.. KOGL_PW=.. python -m api.module.dataset_builder.backfill_to_target \
      --target 500 --types 영상,어문,이미지 --out /mnt/e/kogl_originals
"""
from __future__ import annotations
import argparse, json, shutil, time
from collections import Counter
from pathlib import Path

from .download_originals import login, download_one, DIVISION
from .build import (read_candidates, download_all, SUBFOLDER, METADATA_COLUMNS,
                    MANIFEST_EXCLUDE, EXTRA_COLUMNS, DEFAULT_XLSX, PROJECT_ROOT, parse_counts)

STATE_DEFAULT = PROJECT_ROOT / "dataset" / "backfill_state.json"


def has_file(out: Path, cls: str, idx: str) -> bool:
    d = out / cls
    return d.exists() and any(f.stat().st_size > 0 for f in d.glob(f"{idx}.*"))


def dl_with_relogin(pg, idx, div, sub, uid, pw):
    """download_one + 세션 만료 시 1회 재로그인 후 재시도(백필은 장시간 단일 세션)."""
    res = download_one(pg, idx, div, sub, 0)
    if res["status"] == "not_authenticated":
        try:
            login(pg, uid, pw); pg.wait_for_timeout(1500)
        except Exception:
            pass
        res = download_one(pg, idx, div, sub, 0)
    return res


def load_manifest(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path); ws = wb.active
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
    wb.close()
    return hdr, rows


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="분류별 다운로드 가능 원본을 목표치까지 백필")
    ap.add_argument("--target", type=int, default=500)
    ap.add_argument("--types", default="영상,어문,이미지")
    ap.add_argument("--out", default="/mnt/e/kogl_originals")
    ap.add_argument("--manifest", default=str(PROJECT_ROOT / "dataset" / "manifest.xlsx"))
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    ap.add_argument("--state", default=str(STATE_DEFAULT))
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--max-attempts-per-type", type=int, default=400,
                    help="유형별 백필 시도 상한(비다운로드 후보 낭비 방어)")
    ap.add_argument("--no-reconcile", action="store_true", help="다운로드만, manifest 재작성 생략")
    args = ap.parse_args(argv)

    import os
    uid, pw = os.environ.get("KOGL_ID"), os.environ.get("KOGL_PW")
    if not uid or not pw:
        print("KOGL_ID / KOGL_PW 필요"); return 2

    out = Path(args.out)
    types = [t.strip() for t in args.types.split(",") if t.strip()]
    target = args.target
    hdr, manifest = load_manifest(Path(args.manifest))
    man_by_cls = {c: [r for r in manifest if r["분류"] == c] for c in types}
    manifest_idx = {str(r["원문인덱스"]) for r in manifest}

    state = json.loads(Path(args.state).read_text()) if Path(args.state).exists() else {"tried": [], "added": {}}
    tried = set(state.get("tried", []))
    added = {c: state.get("added", {}).get(c, []) for c in types}  # list of idx added (downloaded)

    from playwright.sync_api import sync_playwright
    with sync_playwright() as p:
        b = p.chromium.launch(headless=True, args=["--no-sandbox"])
        ctx = b.new_context(accept_downloads=True, ignore_https_errors=True,
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148 Safari/537.36")
        pg = ctx.new_page()
        print("로그인 중...", flush=True); login(pg, uid, pw); pg.wait_for_timeout(2000)
        # 로그인 확인
        try:
            pg.goto("https://www.kogl.or.kr/", wait_until="domcontentloaded", timeout=30000)
            print("로그인 OK" if (("로그아웃" in pg.content()) or ("마이페이지" in pg.content())) else "⚠ 로그인 확인 실패", flush=True)
        except Exception:
            pass

        # ---- Phase 1: RETRY non-downloaded manifest works (catch transient) ----
        for cls in types:
            miss = [r for r in man_by_cls[cls] if not has_file(out, cls, str(r["원문인덱스"]))]
            print(f"\n[RETRY] {cls}: 미다운로드 {len(miss)}건 재시도", flush=True)
            for r in miss:
                idx = str(r["원문인덱스"])
                sub = out / cls; sub.mkdir(parents=True, exist_ok=True)
                res = dl_with_relogin(pg, idx, DIVISION.get(cls, "img"), sub, uid, pw)
                if res["status"] == "ok":
                    print(f"  ✅ 회수 {cls} {idx} {res.get('size',0)//1024}KB", flush=True)

        # ---- Phase 2: BACKFILL from export until target downloadable ----
        cand_groups, _ = read_candidates(Path(args.xlsx), set(types))
        backfill_recs = {c: [] for c in types}  # recs (export dicts) successfully downloaded
        for cls in types:
            have = sum(1 for r in man_by_cls[cls] if has_file(out, cls, str(r["원문인덱스"]))) + len(added[cls])
            short = max(0, target - have)
            print(f"\n[BACKFILL] {cls}: 다운로드됨 {have}/{target} → 부족 {short}", flush=True)
            if short == 0:
                continue
            pool = [c for c in cand_groups.get(cls, [])
                    if c["원문인덱스"] not in manifest_idx and c["원문인덱스"] not in tried]
            # 시드 순서 고정(완성도 선호) — build.select_for_class 와 동일 정신, 단 전수 순회
            import random
            rng = random.Random(args.seed); rng.shuffle(pool)
            from .build import completeness_score
            pool.sort(key=completeness_score, reverse=True)
            got = 0; attempts = 0
            for rec in pool:
                if got >= short or attempts >= args.max_attempts_per_type:
                    break
                idx = rec["원문인덱스"]; attempts += 1
                sub = out / cls; sub.mkdir(parents=True, exist_ok=True)
                res = dl_with_relogin(pg, idx, DIVISION.get(cls, "img"), sub, uid, pw)
                tried.add(idx)
                if res["status"] == "ok":
                    got += 1; backfill_recs[cls].append(rec); added[cls].append(idx)
                    print(f"  ✅ 백필 {cls} {idx} ({got}/{short}) {res.get('size',0)//1024}KB", flush=True)
                else:
                    if attempts % 20 == 0:
                        print(f"    …시도 {attempts}, 채택 {got}/{short} (최근 {res['status']})", flush=True)
                if attempts % 10 == 0:
                    Path(args.state).write_text(json.dumps({"tried": sorted(tried),
                        "added": {c: added[c] for c in types}}, ensure_ascii=False))
                time.sleep(0.5)
            print(f"  → {cls} 백필 결과: +{got} (시도 {attempts})", flush=True)
        b.close()
    Path(args.state).write_text(json.dumps({"tried": sorted(tried),
        "added": {c: added[c] for c in types}}, ensure_ascii=False))

    # ---- Phase 3: download backfill thumbnails + RECONCILE manifest ----
    # 재개 안전성: 이번 실행분(backfill_recs)뿐 아니라 state 의 모든 added idx 를 후보맵에서
    # rec 로 복원해 manifest 에 반영한다(이전 실행에서 받은 백필 작품 누락 방지).
    idx2rec = {c["원문인덱스"]: c for cls in types for c in cand_groups.get(cls, [])}
    backfill_recs = {c: [idx2rec[i] for i in added[c] if i in idx2rec] for c in types}
    all_backfill = {c: backfill_recs[c] for c in types}
    if any(all_backfill.values()):
        print("\n[THUMBS] 백필 작품 썸네일 다운로드...", flush=True)
        download_all(all_backfill, PROJECT_ROOT / "dataset", workers=5, seed=args.seed)

    if args.no_reconcile:
        print("\n--no-reconcile: manifest 재작성 생략. 백필 다운로드 완료.", flush=True)
        return 0

    # 새 manifest: 분류별 = 다운로드된 원본(원본+백필). 비다운로드 제외.
    import openpyxl
    manifest_meta_cols = [c for c in METADATA_COLUMNS if c not in MANIFEST_EXCLUDE]
    columns = manifest_meta_cols + EXTRA_COLUMNS
    assert columns == hdr, f"컬럼 불일치 cols={columns} hdr={hdr}"
    # 비-types 분류 행은 그대로 보존
    other_rows = [r for r in manifest if r["분류"] not in types]
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "manifest"
    ws.append(columns)
    for r in other_rows:
        ws.append([r.get(c, "") for c in columns])
    final_counts = Counter()
    for cls in types:
        kept = [r for r in man_by_cls[cls] if has_file(out, cls, str(r["원문인덱스"]))]
        for r in kept:
            ws.append([r.get(c, "") for c in columns]); final_counts[cls] += 1
        for rec in backfill_recs[cls]:
            row = [rec.get(k, "") for k in manifest_meta_cols]
            rel = f"{SUBFOLDER.get(cls,'misc')}/{rec['원문인덱스']}{__import__('pathlib').Path(rec.get('쎔네일웹경로','') or '.jpg').suffix or '.jpg'}"
            row += [rel, "ok", "True", SUBFOLDER.get(cls, "misc")]
            ws.append(row); final_counts[cls] += 1
    backup = Path(args.manifest).with_name("manifest_pre_backfill.xlsx")
    if not backup.exists():
        shutil.copy(args.manifest, backup); print(f"매니페스트 백업: {backup}", flush=True)
    wb.save(args.manifest)
    print(f"\n새 manifest: {dict(final_counts)} (+ 기타 {len(other_rows)}) = {sum(final_counts.values())+len(other_rows)}행", flush=True)
    print("다음: generate_contracts 재실행 + originals_index 재빌드 필요.", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

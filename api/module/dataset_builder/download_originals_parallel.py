"""
KOGL 원본 병렬 다운로더 — download_originals.py 의 단일 세션 흐름을 async 로 포팅하여
**하나의 인증 세션을 N개 워커 페이지가 공유**하며 동시에 내려받는다.

안티-차단 설계:
- 로그인은 1회만(공유 컨텍스트 쿠키). 워커별 로그인 금지(세션 상호 축출 위험).
- 동시성은 보수적으로(기본 3). 워커별 지터(1~3s) 지연으로 버스트 방지.
- 실패가 몰리면(레이트리밋 신호) 전역 백오프 후 재개.
- 세션 만료 시 1회 조정 재로그인(공유 컨텍스트라 전 워커에 반영).
- skip-existing 으로 재개 가능. download_log.json 체크포인트.

사용:
    KOGL_ID=.. KOGL_PW=.. python -m api.module.dataset_builder.download_originals_parallel \
        --types 영상,어문 --dry-run 6 --workers 3 --out /mnt/e/kogl_originals      # 프로브
    KOGL_ID=.. KOGL_PW=.. python -m api.module.dataset_builder.download_originals_parallel \
        --types 이미지,어문,영상 --workers 3 --video-max-mb 0 --out /mnt/e/kogl_originals  # 전체(무제한)
"""
from __future__ import annotations
import argparse, asyncio, json, os, random, time
from collections import Counter
from pathlib import Path

from .download_originals import DIVISION, OPENER, FILL_JS, PURPOSE, DETAIL, ENTRY, DEFAULT_MANIFEST, DEFAULT_OUT


async def login_async(page, uid: str, pw: str) -> None:
    await page.goto(ENTRY, wait_until="domcontentloaded", timeout=45000)
    await page.wait_for_timeout(4500)
    await page.get_by_text("아이디 로그인", exact=False).first.click(timeout=15000)
    await page.wait_for_timeout(2500)
    await page.fill("#id", uid)
    await page.fill("#pw", pw)
    await page.locator("#btn-login").click()
    for _ in range(3):
        try:
            await page.wait_for_load_state("networkidle", timeout=20000)
        except Exception:
            pass
    await page.wait_for_timeout(1500)


async def attempt_async(page, idx: str, division: str, out_dir: Path, video_cap: int) -> dict:
    """download_originals._attempt 의 async 포팅. video_cap=0 이면 무제한."""
    await page.goto(DETAIL.format(idx=idx, div=division), wait_until="domcontentloaded", timeout=45000)
    await page.wait_for_timeout(1500)
    html = await page.content()
    if not (("로그아웃" in html) or ("마이페이지" in html)):
        return {"status": "not_authenticated"}
    opener = page.locator(OPENER)
    if await opener.count() == 0:
        return {"status": "no_in_site_download"}
    await opener.first.click(timeout=10000)
    await page.wait_for_timeout(1800)
    await page.evaluate(FILL_JS, PURPOSE)
    await page.wait_for_timeout(400)
    alert: dict = {}
    clen: dict = {}

    def on_resp(r):
        try:
            cd = r.headers.get("content-disposition", "")
            if ("attachment" in cd) or ("recommFileDown" in r.url):
                cl = r.headers.get("content-length")
                if cl and cl.isdigit():
                    clen["n"] = int(cl)
        except Exception:
            pass

    def on_dlg(d):
        alert["m"] = d.message
        asyncio.create_task(d.accept())

    page.on("dialog", on_dlg)
    page.on("response", on_resp)
    try:
        async with page.expect_download(timeout=30000) as di:
            sub = page.locator("button[onclick*='recomDownloadFile']")
            if await sub.count():
                await sub.first.click(timeout=8000)
            else:
                await page.locator("button[type=submit]:has-text('다운로드'), button:has-text('다운로드')").last.click(timeout=8000)
        d = await di.value
        await page.wait_for_timeout(300)
        if video_cap and division == "video" and clen.get("n", 0) > video_cap:
            try:
                await d.cancel()
            except Exception:
                pass
            return {"status": "skipped_video_large", "size": clen["n"]}
        ext = os.path.splitext(d.suggested_filename or "")[1] or ".bin"
        path = out_dir / f"{idx}{ext}"
        await d.save_as(str(path))
        size = path.stat().st_size
        if video_cap and division == "video" and size > video_cap:
            path.unlink(missing_ok=True)
            return {"status": "skipped_video_large", "size": size}
        return {"status": "ok", "file": path.name, "size": size}
    except Exception as e:
        return {"status": "retry", "error": (alert.get("m") or str(e))[:160]}
    finally:
        page.remove_listener("dialog", on_dlg)
        page.remove_listener("response", on_resp)


async def download_one_async(page, idx, division, out_dir: Path, video_cap, retries=3) -> dict:
    res = {"idx": idx, "division": division, "status": "", "file": "", "size": 0}
    existing = list(out_dir.glob(f"{idx}.*"))
    if existing and existing[0].stat().st_size > 0:
        res.update(status="skipped_existing", file=existing[0].name, size=existing[0].stat().st_size)
        return res
    last_err = last_status = ""
    for attempt in range(1, retries + 1):
        try:
            r = await attempt_async(page, idx, division, out_dir, video_cap)
        except Exception as e:
            r = {"status": "retry", "error": str(e)[:160]}
        st = r["status"]
        if st in ("ok", "skipped_video_large", "no_in_site_download"):
            res.update(r); return res
        last_status, last_err = st, r.get("error", st)
        if attempt < retries:
            await asyncio.sleep(1.5)
    res.update(status=("not_authenticated" if last_status == "not_authenticated" else "failed"), error=last_err)
    return res


MARK = {"ok": "✅", "skipped_existing": "·", "skipped_video_large": "⏭",
        "no_in_site_download": "🚫", "failed": "✗", "not_authenticated": "🔒"}


async def worker(wid, ctx, queue, out: Path, video_cap, results, ctl, args, uid, pw):
    page = await ctx.new_page()
    try:
        while True:
            item = await queue.get()
            if item is None:
                queue.task_done(); return
            try:
                cls, idx, title = item
                div = DIVISION.get(cls, "img")
                sub = out / cls; sub.mkdir(parents=True, exist_ok=True)
                # 전역 백오프 대기
                async with ctl["lock"]:
                    wait = ctl["pause_until"] - time.time()
                if wait > 0:
                    await asyncio.sleep(wait)
                res = await download_one_async(page, idx, div, sub, video_cap)
                # 세션 만료 → 조정된 1회 재로그인(공유 컨텍스트 쿠키 갱신 → 전 워커 반영)
                if res["status"] == "not_authenticated":
                    async with ctl["relogin_lock"]:
                        try:
                            await login_async(page, uid, pw); await page.wait_for_timeout(1000)
                        except Exception:
                            pass
                    res = await download_one_async(page, idx, div, sub, video_cap)
                res["분류"] = cls; res["제목"] = title
                async with ctl["lock"]:
                    results.append(res); n = len(results)
                    if res["status"] in ("failed", "not_authenticated"):
                        ctl["fails"] += 1
                        if ctl["fails"] >= args.backoff_after:
                            ctl["pause_until"] = time.time() + args.backoff_secs
                            ctl["fails"] = 0
                            print(f"  ⏸ 백오프 {args.backoff_secs}s (실패 누적)", flush=True)
                    elif res["status"] == "ok":
                        ctl["fails"] = 0
                    if n % args.checkpoint == 0:
                        (out / "download_log.json").write_text(json.dumps(results, ensure_ascii=False))
                print(f"  {MARK.get(res['status'],'?')} w{wid} [{n}/{ctl['total']}] {cls} {idx} {res['status']} {res.get('size',0)//1024}KB", flush=True)
                await asyncio.sleep(random.uniform(args.jitter_min, args.jitter_max))
            except Exception as e:
                print(f"  ! w{wid} {item} {str(e)[:80]}", flush=True)
            finally:
                queue.task_done()
    finally:
        try:
            await page.close()
        except Exception:
            pass


async def main_async(args) -> int:
    uid, pw = os.environ.get("KOGL_ID"), os.environ.get("KOGL_PW")
    if not uid or not pw:
        print("KOGL_ID / KOGL_PW 환경변수가 필요합니다 (.env 권장)."); return 2
    import openpyxl
    ws = openpyxl.load_workbook(args.manifest).active
    hdr = [c.value for c in ws[1]]; ci = {h: i for i, h in enumerate(hdr)}
    types = [t.strip() for t in args.types.split(",") if t.strip()]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[ci["분류"]] in types]
    if args.dry_run:
        sel = []
        for t in types:
            sel += [r for r in rows if r[ci["분류"]] == t][: args.dry_run]
        rows = sel
    elif args.limit:
        rows = rows[: args.limit]
    work = [(r[ci["분류"]], str(r[ci["원문인덱스"]]), r[ci["제목"]]) for r in rows]
    # 이미지(소·빠름) → 어문 → 영상(대·느림) 순서: 빠른 성과 우선 + 중단 시 회수 최대화
    _prio = {"이미지": 0, "어문": 1, "영상": 2}
    work.sort(key=lambda w: _prio.get(w[0], 9))
    out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
    video_cap = args.video_max_mb * 1024 * 1024 if args.video_max_mb > 0 else 0
    print(f"대상 {len(work)}건 {dict(Counter(w[0] for w in work))} | workers={args.workers} | "
          f"영상캡={'무제한' if not video_cap else str(args.video_max_mb)+'MB'} | 지터 {args.jitter_min}-{args.jitter_max}s | 출력 {out}", flush=True)

    queue: asyncio.Queue = asyncio.Queue()
    for w in work:
        queue.put_nowait(w)
    for _ in range(args.workers):
        queue.put_nowait(None)
    results: list = []
    ctl = {"lock": asyncio.Lock(), "relogin_lock": asyncio.Lock(),
           "pause_until": 0.0, "fails": 0, "total": len(work)}
    t0 = time.time()
    from playwright.async_api import async_playwright
    async with async_playwright() as p:
        b = await p.chromium.launch(headless=True, args=["--no-sandbox"])
        ctx = await b.new_context(accept_downloads=True, ignore_https_errors=True,
            user_agent="Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148 Safari/537.36")
        boot = await ctx.new_page()
        print("로그인 중...", flush=True)
        await login_async(boot, uid, pw)
        await boot.wait_for_timeout(2000)
        ok = False
        for _ in range(3):
            try:
                await boot.goto("https://www.kogl.or.kr/", wait_until="domcontentloaded", timeout=30000)
                c = await boot.content(); ok = ("로그아웃" in c) or ("마이페이지" in c); break
            except Exception:
                await boot.wait_for_timeout(2000)
        print("로그인 OK" if ok else "⚠ 로그인 확인 실패", flush=True)
        await boot.close()
        tasks = [asyncio.create_task(worker(i + 1, ctx, queue, out, video_cap, results, ctl, args, uid, pw))
                 for i in range(args.workers)]
        await asyncio.gather(*tasks)
        await b.close()

    dt = time.time() - t0
    dl = sum(r.get("size", 0) for r in results if r["status"] == "ok")
    (out / "download_log.json").write_text(json.dumps(results, ensure_ascii=False))
    print(f"\n=== 요약 === {dict(Counter(r['status'] for r in results))}", flush=True)
    print(f"  ok 다운로드 {dl/1024**3:.2f}GB / {dt/60:.1f}분 => 집계 {dl/1048576/max(dt,1):.2f} MB/s | 로그 {out}/download_log.json", flush=True)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="KOGL 원본 병렬 다운로더 (단일 세션 공유, Playwright async)")
    ap.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--types", default="이미지,어문,영상")
    ap.add_argument("--workers", type=int, default=3, help="동시 워커(페이지) 수 — 차단 방지 위해 보수적으로(2~4 권장)")
    ap.add_argument("--dry-run", type=int, default=0, help="유형별 N건만")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--video-max-mb", type=int, default=0, help="영상 용량 상한(MB). 0=무제한")
    ap.add_argument("--jitter-min", type=float, default=1.0)
    ap.add_argument("--jitter-max", type=float, default=3.0)
    ap.add_argument("--backoff-after", type=int, default=6, help="연속 실패 N회 시 백오프")
    ap.add_argument("--backoff-secs", type=int, default=60)
    ap.add_argument("--checkpoint", type=int, default=25)
    args = ap.parse_args()
    return asyncio.run(main_async(args))


if __name__ == "__main__":
    raise SystemExit(main())

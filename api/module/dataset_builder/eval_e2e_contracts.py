"""
End-to-end 파이프라인 평가 — 합성 계약서(PDF) N건을 실제 배포 파이프라인
(OCR → LLM ∥ NER → consolidation)으로 처리하고 contracts_index.xlsx 정답과 대조.

Paper 2 (IC-EEECS) Tables I/II + latency + (LLM-only vs consolidated) ablation 산출.
저널 §8 end-to-end eval 의 시드이기도 함.

채점 설계:
  Tier-1 (schema-anchored, strict-normalized field match):
      work_title↔제목, copyright_holder↔저작권자명, kogl_type↔공공누리 유형
  Tier-2 (value recall — 정답 값이 추출 결과 어디든 정규화 포함되는가):
      당사자 PII 13필드 (갑 담당자/연락처/이메일/사업자번호, 을 식별·연락 필드들)
      LLM-only(metadata) vs consolidated(consolidated_metadata) 각각 채점 → ablation.

재개 가능: results.jsonl 에 idx 단위 체크포인트, 이미 있는 idx 는 스킵.

사용:
  python -m api.module.dataset_builder.eval_e2e_contracts --limit 2          # pilot
  python -m api.module.dataset_builder.eval_e2e_contracts                    # full sample
  python -m api.module.dataset_builder.eval_e2e_contracts --report-only      # 집계만
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "api"))

PDF_DIR = ROOT / "dataset" / "contracts_pdf"
GT_XLSX = ROOT / "dataset" / "contracts" / "contracts_index.xlsx"
OUT_DIR = ROOT / "dataset" / "e2e_eval"

# Tier-1: 정답컬럼 -> 통합 스키마 필드 (strict field match)
TIER1 = {"제목": "work_title", "저작권자명": "copyright_holder", "공공누리 유형": "kogl_type"}
# Tier-2: 정답컬럼 -> (값 종류) — value recall anywhere in extracted output
TIER2 = {
    "갑_담당자": "name", "갑_연락처": "phone", "갑_이메일": "email", "갑_사업자등록번호": "digits",
    "이용자명": "name", "을_대표자": "name", "을_사업자등록번호": "digits",
    "을_주민등록번호": "digits", "을_생년월일": "digits", "을_휴대폰": "phone",
    "을_전화": "phone", "을_이메일": "email", "을_주소": "text",
}
GROUPS = {  # Paper Table I rows
    "work": ["제목", "저작권자명", "공공누리 유형"],
    "party": ["갑_담당자", "이용자명", "을_대표자"],
    "contact": ["갑_연락처", "갑_이메일", "을_휴대폰", "을_전화", "을_이메일", "을_주소"],
    "identifier": ["갑_사업자등록번호", "을_사업자등록번호", "을_주민등록번호", "을_생년월일"],
}


def _norm_text(s) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return re.sub(r"\s+", "", s).lower()


def _digits(s) -> str:
    return re.sub(r"\D", "", str(s or ""))


def _norm_kogl(v) -> str:
    """공공누리 유형 정규화: '제1유형'/'1'/'공공누리 제1유형' → '1'."""
    m = re.search(r"[0-4]", str(v or ""))
    return m.group(0) if m else ""


def _value_in_blob(gt_value, kind: str, blob_text: str, blob_digits: str) -> bool:
    if gt_value in (None, "", "None"):
        return None  # not applicable
    if kind in ("digits", "phone"):
        d = _digits(gt_value)
        return bool(d) and d in blob_digits
    if kind == "email":
        return _norm_text(gt_value) in blob_text
    # name/text: normalized containment
    n = _norm_text(gt_value)
    return bool(n) and n in blob_text


def _blobs(meta: dict) -> tuple[str, str]:
    txt = json.dumps(meta, ensure_ascii=False) if meta else ""
    return _norm_text(txt), _digits(txt)


def score_doc(gt: dict, llm_meta: dict, con_meta: dict) -> dict:
    out = {"tier1": {}, "tier2": {}}
    for src in ("llm", "con"):
        meta = llm_meta if src == "llm" else con_meta
        meta = meta or {}
        bt, bd = _blobs(meta)
        # Tier-1 strict field
        for col, field in TIER1.items():
            gtv = gt.get(col)
            if gtv in (None, "", "None"):
                out["tier1"].setdefault(col, {})[src] = None
                continue
            ext = meta.get(field)
            if col == "공공누리 유형":
                ok = _norm_kogl(ext) == _norm_kogl(gtv) and _norm_kogl(gtv) != ""
            else:
                ok = bool(_norm_text(ext)) and (_norm_text(gtv) in _norm_text(ext) or _norm_text(ext) in _norm_text(gtv))
            out["tier1"].setdefault(col, {})[src] = bool(ok)
        # Tier-2 value recall
        for col, kind in TIER2.items():
            out["tier2"].setdefault(col, {})[src] = _value_in_blob(gt.get(col), kind, bt, bd)
    return out


def load_gt() -> dict:
    import openpyxl
    ws = openpyxl.load_workbook(GT_XLSX).active
    hdr = [c.value for c in ws[1]]
    return {str(r[0]): dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)}


def build_orchestrator():
    from api import ner_predict  # noqa: heavy import (torch)
    from module.llm_extraction import LLMExtractionProcessor
    from web.pipeline import PipelineOrchestrator
    out = OUT_DIR / "runs"
    out.mkdir(parents=True, exist_ok=True)
    llm_processor = LLMExtractionProcessor(output_dir=str(out / "llm_results"))
    models = {"klue-roberta-large": {"name": "klue/roberta-large", "display_name": "KLUE RoBERTa Large"}}
    return PipelineOrchestrator(
        llm_processor=llm_processor, ner_predict_fn=ner_predict,
        available_ner_models=models, upload_dir=out / "uploads", results_dir=out / "results")


def run_one(orch, idx: str, pdf: Path, args) -> dict:
    rec = {"idx": idx, "ok": False}
    t0 = time.perf_counter()
    ctx = orch.setup(pdf.read_bytes(), f"{idx}.pdf")
    t1 = time.perf_counter()
    ocr_text, ocr_result = orch.run_ocr(ctx, args.ocr_provider, args.ocr_model)
    t2 = time.perf_counter()
    rec["ocr_s"] = round(t2 - t1, 1)
    rec["ocr_chars"] = len(ocr_text)
    if not ocr_text.strip():
        rec["error"] = "empty OCR"
        return rec
    tl = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        def _llm():
            s = time.perf_counter()
            r = orch.run_llm(ocr_text, args.document_type, ctx["filename"], args.llm_model)
            tl["llm"] = time.perf_counter() - s
            return r
        def _ner():
            s = time.perf_counter()
            r = orch.run_ner(ocr_result, ctx["result_dir"], "klue-roberta-large", ocr_text)
            tl["ner"] = time.perf_counter() - s
            return r
        lf, nf = ex.submit(_llm), ex.submit(_ner)
        llm_result, ner_result = lf.result(), nf.result()
    rec["llm_s"] = round(tl.get("llm", 0), 1)
    rec["ner_s"] = round(tl.get("ner", 0), 1)
    rec["llm_ok"] = bool(llm_result.get("success"))
    rec["ner_ok"] = bool(ner_result.get("success"))
    rec["ner_entities"] = ner_result.get("total_entities", 0)
    t3 = time.perf_counter()
    con_result, con_ok, con_err = orch.run_consolidation(
        llm_result, ner_result, ocr_text, args.document_type,
        ctx["result_dir"], args.consolidation_model)
    rec["con_s"] = round(time.perf_counter() - t3, 1)
    rec["con_ok"] = bool(con_ok)
    rec["total_s"] = round(time.perf_counter() - t0, 1)

    llm_meta = llm_result.get("metadata") or {}
    con_meta = (con_result or {}).get("consolidated_metadata") or {}
    rec["llm_meta"] = llm_meta
    rec["con_meta"] = con_meta
    vr = (con_result or {}).get("validation_report") or {}
    decisions = vr.get("decisions") or []
    rec["decisions"] = [{"field": d.get("field") or d.get("field_name"),
                         "decision": d.get("decision") or d.get("status"),
                         "confidence": d.get("confidence")} for d in decisions]
    rec["con_confidence"] = vr.get("confidence_score")
    rec["ok"] = True
    return rec


def aggregate(results: list, out_md: Path) -> str:
    import statistics as st
    done = [r for r in results if r.get("ok")]
    L = [f"# E2E pipeline eval — {len(done)}/{len(results)} docs OK\n"]
    # Table I: per group, llm vs con
    L.append("## Table I — field accuracy / value recall (%) [LLM-only vs Consolidated]")
    L.append("| Group | n | LLM-only | Consolidated |")
    L.append("|---|---|---|---|")
    for gname, cols in GROUPS.items():
        for src in ():
            pass
        stats = {"llm": [0, 0], "con": [0, 0]}
        for r in done:
            sc = r.get("score", {})
            for col in cols:
                tier = "tier1" if col in TIER1 else "tier2"
                v = sc.get(tier, {}).get(col, {})
                for src in ("llm", "con"):
                    val = v.get(src)
                    if val is None:
                        continue
                    stats[src][1] += 1
                    stats[src][0] += 1 if val else 0
        row = [gname, str(stats["llm"][1])]
        for src in ("llm", "con"):
            c, n = stats[src]
            row.append(f"{100*c/n:.1f}" if n else "–")
        L.append("| " + " | ".join(row) + " |")
    # overall
    tot = {"llm": [0, 0], "con": [0, 0]}
    for r in done:
        sc = r.get("score", {})
        for tier in ("tier1", "tier2"):
            for col, v in sc.get(tier, {}).items():
                for src in ("llm", "con"):
                    val = v.get(src)
                    if val is None:
                        continue
                    tot[src][1] += 1
                    tot[src][0] += 1 if val else 0
    L.append(f"| **Overall** | {tot['llm'][1]} | **{100*tot['llm'][0]/max(tot['llm'][1],1):.1f}** | **{100*tot['con'][0]/max(tot['con'][1],1):.1f}** |")
    # Table II decisions
    from collections import Counter, defaultdict
    dc = Counter(); dconf = defaultdict(list)
    for r in done:
        for d in r.get("decisions", []):
            k = (d.get("decision") or "?").upper()
            dc[k] += 1
            if isinstance(d.get("confidence"), (int, float)):
                dconf[k].append(d["confidence"])
    total_d = sum(dc.values()) or 1
    L.append("\n## Table II — consolidation decision distribution")
    L.append("| Decision | share | mean confidence |")
    L.append("|---|---|---|")
    for k, n in dc.most_common():
        mc = f"{st.mean(dconf[k]):.2f}" if dconf[k] else "–"
        L.append(f"| {k} | {100*n/total_d:.1f}% | {mc} |")
    # latency
    L.append("\n## Latency (median seconds/doc)")
    for k in ("ocr_s", "llm_s", "ner_s", "con_s", "total_s"):
        vals = [r[k] for r in done if isinstance(r.get(k), (int, float))]
        if vals:
            L.append(f"- {k}: median {st.median(vals):.1f}s (mean {st.mean(vals):.1f}s)")
    par = [max(r.get("llm_s", 0), r.get("ner_s", 0)) - r.get("ner_s", 0) for r in done]
    seq_save = [min(r.get("llm_s", 0), r.get("ner_s", 0)) for r in done]
    if seq_save:
        L.append(f"- LLM∥NER concurrency saves ≈ median {st.median(seq_save):.1f}s/doc vs sequential")
    md = "\n".join(L)
    out_md.write_text(md, encoding="utf-8")
    return md


def main() -> int:
    ap = argparse.ArgumentParser(description="E2E contracts pipeline eval")
    ap.add_argument("--sample", default=str(PDF_DIR / "sample100.json"))
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--ocr-provider", default="alibaba")
    ap.add_argument("--ocr-model", default=None)
    ap.add_argument("--llm-model", default="alibaba-qwen3.5-122b-a10b")
    ap.add_argument("--consolidation-model", default="alibaba-qwen3.5-122b-a10b")
    ap.add_argument("--document-type", default="계약서")
    ap.add_argument("--report-only", action="store_true")
    args = ap.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    res_path = OUT_DIR / "results.jsonl"
    results = []
    if res_path.exists():
        results = [json.loads(l) for l in res_path.read_text(encoding="utf-8").splitlines() if l.strip()]
    done_idx = {r["idx"] for r in results}

    if not args.report_only:
        sample = json.load(open(args.sample))
        if args.limit:
            sample = sample[: args.limit]
        todo = [i for i in sample if i not in done_idx]
        print(f"sample={len(sample)} done={len(done_idx & set(sample))} todo={len(todo)}", flush=True)
        gt_all = load_gt()
        if todo:
            print("파이프라인 부트스트랩 (torch/NER 로드, ~1분)...", flush=True)
            orch = build_orchestrator()
            for i, idx in enumerate(todo, 1):
                pdf = PDF_DIR / f"{idx}.pdf"
                if not pdf.exists():
                    print(f"  ✗ [{i}/{len(todo)}] {idx}: no PDF", flush=True)
                    continue
                try:
                    rec = run_one(orch, idx, pdf, args)
                except Exception as e:
                    rec = {"idx": idx, "ok": False, "error": str(e)[:200]}
                if rec.get("ok"):
                    rec["score"] = score_doc(gt_all.get(idx, {}), rec.get("llm_meta", {}), rec.get("con_meta", {}))
                # keep JSONL small: drop bulky metadata after scoring
                rec.pop("llm_meta", None); rec.pop("con_meta", None)
                results.append(rec)
                with open(res_path, "a", encoding="utf-8") as f:
                    f.write(json.dumps(rec, ensure_ascii=False) + "\n")
                mark = "✅" if rec.get("ok") else "✗"
                print(f"  {mark} [{i}/{len(todo)}] {idx} total={rec.get('total_s','?')}s "
                      f"ocr={rec.get('ocr_s','?')} llm={rec.get('llm_s','?')} ner={rec.get('ner_s','?')} "
                      f"con={rec.get('con_s','?')} {rec.get('error','')}", flush=True)

    md = aggregate(results, OUT_DIR / "report.md")
    print("\n" + md, flush=True)
    print(f"\n[written] {OUT_DIR/'report.md'} | raw: {res_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

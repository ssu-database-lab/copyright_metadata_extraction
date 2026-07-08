"""
계약서 생성기 — manifest.xlsx(실제 저작물 메타데이터) + 합성 PII → 저작재산권 이용허락
계약서 .docx (저작물 1건당 1부).

흐름: 템플릿(검증된 markdown) 채우기 → pandoc로 .docx 변환.
- 갑(저작권자): 매니페스트의 실제 저작권자명/소속 + 합성 담당자 연락처.
- 을(이용자): 완전 합성(개인/법인) PII.
- 값 처리 규칙(검증 완료): (1) '-'/None/'' = 빈값, (2) 유효기간 3분기(실제일자/만료저작물/
  boilerplate), (3) 저작권자명 generic('개인' 등) → "(성명 미상의 개인 저작권자)".
- 부산물: contracts_index.xlsx — 각 계약서에 넣은 합성 PII + 실제 메타 = 추출 평가용 정답.

사용:
    python -m api.module.dataset_builder.generate_contracts            # 전체(manifest 1000건)
    python -m api.module.dataset_builder.generate_contracts --dry-run 3
    python -m api.module.dataset_builder.generate_contracts --workers 6 --out dataset/contracts
"""

from __future__ import annotations

import argparse
import re
import subprocess
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from api.module.dataset_builder.synthetic_identity import (
    make_licensee, make_licensor_contact, is_institution,
)

ROOT = Path(__file__).resolve().parents[3]
TEMPLATE = Path(__file__).resolve().parent / "templates" / "저작재산권_이용허락_계약서_template.md"
DEFAULT_MANIFEST = ROOT / "dataset" / "manifest.xlsx"
DEFAULT_OUT = ROOT / "dataset" / "contracts"

_EMPTY = {"", "none", "-", "--", "n/a", "해당없음", "미상", "불명"}
_GENERIC_OWNERS = {"개인", "단체", "기관", "미상", "불명"}
RIGHTS = [("복제권", "복제권"), ("저작재산권 공연권", "공연권"), ("공중송신권", "공중송신권"),
          ("저작재산권 전시권", "전시권"), ("저작재산권 배포권", "배포권"),
          ("대여권", "대여권"), ("2차적저작물 작성권", "2차적저작물작성권")]


def cb(v) -> str:
    """clean-or-blank: '-'/None/'' 등 placeholder → '' """
    s = "" if v is None else str(v).strip()
    return "" if s.lower() in _EMPTY else s


def _date(s: str) -> str:
    """'2092-12-20 00:00:00' → '2092-12-20'; '2084' 그대로."""
    return s.split(" ")[0].strip()


def computed(row: dict) -> dict:
    code = cb(row.get("공공누리 유형")) or "0"
    granted = [disp for col, disp in RIGHTS if (row.get(col) or "").strip() == "Y"]
    rng = "\n".join(f"   {i}. {x}" for i, x in enumerate(granted, 1)) or "   별도 협의에 따른다."

    com = cb(row.get("상업적이용허락"))
    com_c = ("상업적 이용을 허락한다." if com == "상업적이용가능"
             else "상업적 이용을 금지한다." if com in ("상업적이용금지", "상업적이용불가")
             else "상업적 이용 여부는 별도 협의에 따른다.")

    sec = (row.get("2차적저작물 작성권") or "").strip()
    sec_c = ("2차적저작물의 작성을 허락한다." if sec == "Y"
             else "2차적저작물의 작성을 허락하지 아니한다." if sec == "N"
             else "별도 협의에 따른다.")

    valid = cb(row.get("계약상 유효기간"))
    exp = cb(row.get("저작권 만료일"))
    if valid:
        vc = f"본 계약의 이용허락 기간은 {valid}으로 한다."
    elif exp in ("만료저작물", "만료"):
        vc = "본 저작물은 저작재산권 보호기간이 만료된 공유저작물(만료저작물)로서 자유롭게 이용할 수 있다."
    elif exp:
        vc = f"본 저작물의 저작재산권 보호기간({_date(exp)}까지) 동안 유효하다."
    else:
        vc = "별도의 정함이 없는 한, 「저작권법」상 저작재산권 보호기간 동안 유효하다."

    inj = cb(row.get("저작인접권자"))
    inj_c = (f"저작인접권자({inj})의 권리가 존재하며, 그 처리 책임은 갑에게 있다." if inj else "해당 사항 없음.")
    por = cb(row.get("초상권"))
    por_c = ("본 저작물에 초상이 포함되어 있어, 촬영 동의 및 초상권 분쟁에 대한 책임은 갑에게 있다."
             if por == "해당" else "해당 사항 없음.")
    return {
        "공공누리유형표시": f"제{code}유형", "이용허락_권리범위": rng,
        "상업적이용_조항": com_c, "2차적저작물_조항": sec_c, "유효기간_조항": vc,
        "저작인접권_조항": inj_c, "초상권_조항": por_c,
    }


def build_subs(row: dict) -> dict:
    idx = int(cb(row.get("원문인덱스")) or 0)
    owner_raw = cb(row.get("저작권자명"))
    owner_disp = "(성명 미상의 개인 저작권자)" if owner_raw in _GENERIC_OWNERS else owner_raw
    subs = {
        "저작권자명": owner_disp,
        "저작권자소속": cb(row.get("저작권자 소속")) or "(소속 미상)",
        "공동저작자": cb(row.get("공동저작자")) or "(해당 없음)",
        "저작물명": cb(row.get("제목")),
        "저작물유형": cb(row.get("정보유형")) or cb(row.get("분류")),
        "식별번호": cb(row.get("원문인덱스")),
        "원본파일명": cb(row.get("원본파일명")) or "(파일명 미상)",
        "언어": cb(row.get("언어")) or "한국어",
        "제작일자": cb(row.get("제작일자")) or "(미상)",
        "공표일자": cb(row.get("공표일자")) or "(미상)",
        "주제어": cb(row.get("주제어")) or "(없음)",
        "원본소유자": cb(row.get("원본소유자")),
        "작성일": ARGS_DATE,
    }
    subs.update(computed(row))
    subs.update(make_licensor_contact(idx, owner_raw, cb(row.get("저작권자 소속"))))
    subs.update(make_licensee(idx))
    return subs


def fill(template: str, subs: dict) -> str:
    out = template
    for k, v in subs.items():
        out = out.replace("{{" + k + "}}", "" if v is None else str(v))
    return out


def md_to_docx(md: str, out_path: Path) -> None:
    with tempfile.NamedTemporaryFile("w", suffix=".md", delete=False, encoding="utf-8") as tf:
        tf.write(md)
        tmp = tf.name
    try:
        subprocess.run(["pandoc", tmp, "-f", "markdown", "-o", str(out_path)],
                       check=True, capture_output=True)
    finally:
        Path(tmp).unlink(missing_ok=True)


# 추출 평가용 정답으로 저장할 합성 PII 필드
_INDEX_FIELDS = ["원문인덱스", "제목", "저작권자명", "공공누리 유형",
                 "갑_담당자", "갑_연락처", "갑_이메일", "갑_사업자등록번호",
                 "을_유형", "이용자명", "을_대표자", "을_사업자등록번호",
                 "을_주민등록번호", "을_생년월일", "을_휴대폰", "을_전화", "을_이메일", "을_주소"]


def main() -> int:
    global ARGS_DATE
    ap = argparse.ArgumentParser(description="저작재산권 이용허락 계약서 .docx 생성기")
    ap.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--dry-run", type=int, default=0, help="N건만 생성(검수용)")
    ap.add_argument("--workers", type=int, default=6)
    ap.add_argument("--date", default="2026-06-15", help="계약 작성일")
    args = ap.parse_args()
    ARGS_DATE = args.date

    import openpyxl
    template = TEMPLATE.read_text(encoding="utf-8")
    ws = openpyxl.load_workbook(args.manifest).active
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, r)) for r in ws.iter_rows(min_row=2, values_only=True)]
    if args.dry_run:
        rows = rows[: args.dry_run]
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"템플릿: {TEMPLATE.name} | 매니페스트: {len(rows)}건 | 출력: {out_dir}")

    index_rows = []
    results = {"ok": 0, "skip": 0, "fail": 0}

    def one(row):
        idx = cb(row.get("원문인덱스"))
        out_path = out_dir / f"{idx}.docx"
        subs = build_subs(row)
        if out_path.exists() and out_path.stat().st_size > 0:
            return ("skip", row, subs)
        try:
            md_to_docx(fill(template, subs), out_path)
            return ("ok", row, subs)
        except Exception as e:  # noqa: BLE001
            return ("fail", row, subs, str(e))

    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = [ex.submit(one, r) for r in rows]
        for i, f in enumerate(as_completed(futs), 1):
            res = f.result()
            status, row, subs = res[0], res[1], res[2]
            results[status] += 1
            merged = {**row, **subs}
            index_rows.append({k: merged.get(k, "") for k in _INDEX_FIELDS})
            if status == "fail":
                print(f"  ✗ idx={cb(row.get('원문인덱스'))}: {res[3]}")
            if i % 100 == 0:
                print(f"  진행 {i}/{len(rows)}")

    # 정답 인덱스 저장
    iwb = openpyxl.Workbook(); iws = iwb.active; iws.title = "contracts_index"
    iws.append(_INDEX_FIELDS)
    for r in index_rows:
        iws.append([r[k] for k in _INDEX_FIELDS])
    iwb.save(out_dir / "contracts_index.xlsx")

    print(f"\n=== 생성 요약 ===\n  ok={results['ok']} skip={results['skip']} fail={results['fail']}")
    print(f"  계약서: {out_dir}/*.docx")
    print(f"  정답 인덱스(합성 PII): {out_dir}/contracts_index.xlsx")
    return 0 if results["fail"] == 0 else 1


ARGS_DATE = "2026-06-15"

if __name__ == "__main__":
    raise SystemExit(main())

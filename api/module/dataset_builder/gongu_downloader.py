"""
공유마당(gongu.copyright.or.kr) 저작물 다운로더 — 익명 HTTP(로그인·Playwright 불필요).

검증된 레시피(2026-07-27 워크플로 전수검증):
  1) 목록 GET listWrt{Image,Video,Text,Sound}.do?menuNo=..&pageIndex=p&pageUnit=N&licenseCd=LC
     → 정규식 view.do?wrtSn=(\\d+) → dedupe. pageUnit=100/500 동작 확인.
  2) 상세 GET view.do?wrtSn=N&menuNo=M → <dl><dt>라벨</dt><dd>값</dd> 파싱
  3) hosted 게이트: 원문제공 dd가 '원문파일명 …' → HOSTED / '원문URL' → EXTERNAL(파일 없음)
  4) 파일 GET wrtFileDownload.do?wrtSn=N&fileSn=k (k=1..N, soft-404 sentinel까지)

★ 모든 실패는 사유(issue)와 함께 기록된다 — 결손 자체가 검증된 산출물이 되도록.
   status: ok / skipped_existing / dry_run / filtered_* / failed
   issue  : external_no_file · no_file_soft404 · http_error · timeout · conn_error
            · size_exceeded · detail_fail · truncated · write_error
   → gongu_index.xlsx(전체) + issues 시트 + records.jsonl(증분, 재개용) + issues.jsonl

Gotchas (전부 실측):
  - 파일명: Content-Disposition 이 %-인코딩(어문 전역·이미지 다수) 또는 raw UTF-8 latin-1 mojibake
    → unquote 우선, 실패 시 latin-1→utf-8.
  - gzip 자동해제(requests). Content-Type 빈 값 정상 → magic byte로 판정.
  - soft-404: HTTP 200 + 정확히 60바이트 "<script>alert('File Not Found[N]');history.back();</script>"
    → 최소 60바이트 확보 후 검사(앞 16바이트만 보면 오탐).
  - Accept-Ranges 없음 → 파일 단위 이어받기 불가. .part 기록 후 완료 시 rename.
  - 영상 평균 433MB(max 7.8GB) → 반드시 stream=True. Content-Length 없음(chunked)
    → 컨테이너 헤더(ASF/MP4/MKV/SWF)로 사전 크기 산출 후 --max-file-mb 초과 시 중단.
  - 첫 페이지 500건은 소수 업로더 편중 → --pages 로 층화 샘플링 필수.

사용:
  # 층화 샘플링 + hosted 게이트
  python -m api.module.dataset_builder.gongu_downloader --menu image --license 97 \
      --page-unit 100 --pages 1,27,53,79,105 --limit 500 --hosted-only \
      --workers 4 --out /mnt/e/gongu_dataset/image/expired
  # 사전 선정된 wrtSn 목록으로(영상 census 재활용)
  python -m api.module.dataset_builder.gongu_downloader --menu video \
      --wrtsn-file sel.txt --limit 500 --max-file-mb 800 --workers 2 --out .../video/expired
"""
from __future__ import annotations
import argparse, hashlib, html, json, os, re, struct, threading, time
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from urllib.parse import unquote

import requests

BASE = "https://gongu.copyright.or.kr"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

MENU = {
    "image": ("listWrtImage", "200018"),
    "video": ("listWrtVideo", "200026"),
    "text":  ("listWrtText",  "200019"),
    "sound": ("listWrtSound", "200020"),
}

LICENSE_NAME = {
    "01": "공공누리 제1유형(출처표시)", "02": "공공누리 제2유형(출처표시+상업금지)",
    "03": "공공누리 제3유형(출처표시+변경금지)", "04": "공공누리 제4유형(출처표시+상업금지+변경금지)",
    "20": "PUBLIC DOMAIN(CC0/PD)", "21": "CC BY", "22": "CC BY-ND", "23": "CC BY-SA",
    "24": "CC BY-NC", "25": "CC BY-NC-ND", "26": "CC BY-NC-SA",
    "97": "자유이용 만료", "98": "기증저작물 자유이용", "99": "기증저작물 이용허락",
}

_MAGIC = [
    (b"\xff\xd8\xff", "jpg"), (b"\x89PNG\r\n\x1a\n", "png"), (b"GIF8", "gif"),
    (b"%PDF", "pdf"), (b"PK\x03\x04", "zip/docx"), (b"RIFF", "riff"),
    (b"\xd0\xcf\x11\xe0", "hwp/ole"), (b"ID3", "mp3"), (b"\xff\xfb", "mp3"),
    (b"\x30\x26\xb2\x75", "wmv/asf"), (b"\x1a\x45\xdf\xa3", "mkv"),
    (b"FWS", "swf"), (b"CWS", "swf"), (b"ZWS", "swf"),
]

_TAG = re.compile(r"<[^>]+>")
_DL = re.compile(r"<dt[^>]*>\s*(.*?)\s*</dt>\s*<dd[^>]*>(.*?)</dd>", re.S)
_WRTSN = re.compile(r"view\.do\?wrtSn=(\d+)")
_LICENSE_IMG = re.compile(r"img_license(\d+)\.png")
_TOTAL = re.compile(r"<strong>([\d,]+)</strong>\s*건")
_FNF = re.compile(r"File\s*Not\s*Found", re.I)
_SENTINEL_MIN = 96          # soft-404 본문은 정확히 60B — 여유 확보 후 검사
ASF_HDR = bytes.fromhex("3026b2758e66cf11a6d900aa0062ce6c")
ASF_DATA = bytes.fromhex("3626b2758e66cf11a6d900aa0062ce6c")


def _clean(s: str) -> str:
    if not s:
        return ""
    s = re.sub(r"<br\s*/?>", "\n", s, flags=re.I)
    s = _TAG.sub(" ", html.unescape(s)).replace("‚", ",")
    lines = [ln.strip() for ln in s.splitlines()]
    return re.sub(r"[ \t]+", " ", "\n".join(l for l in lines if l)).strip()


# ── thread-local sessions (requests.Session 은 스레드 공유 비권장) ──────────
_tls = threading.local()


def sess() -> requests.Session:
    s = getattr(_tls, "s", None)
    if s is None:
        s = requests.Session()
        s.headers.update({"User-Agent": UA, "Referer": BASE + "/"})
        _tls.s = s
    return s


# ── 컨테이너 헤더로 전체 크기 산출 (실측 5/5 정확) ─────────────────────────
def size_from_buf(b: bytes) -> Tuple[str, Optional[int]]:
    """앞부분 바이트만으로 (포맷, 총바이트|None) 추정."""
    if len(b) < 16:
        return ("?", None)
    if b.startswith(ASF_HDR):
        i = b.find(ASF_DATA, 16)
        if i >= 0 and len(b) >= i + 24:
            return ("asf/wmv", i + struct.unpack_from("<Q", b, i + 16)[0])
        return ("asf/wmv", None)
    if b[:3] in (b"FWS", b"CWS", b"ZWS"):
        return ("swf", struct.unpack_from("<I", b, 4)[0])
    if b.startswith(bytes.fromhex("1a45dfa3")):
        i = b.find(bytes.fromhex("18538067"))
        if i >= 0 and len(b) > i + 12:
            p = i + 4
            f = b[p]
            if f == 0:
                return ("mkv", None)
            ln = 8 - f.bit_length() + 1
            raw = b[p:p + ln]
            val = raw[0] & ((1 << (8 - ln)) - 1)
            for c in raw[1:]:
                val = (val << 8) | c
            if val == (1 << (7 * ln)) - 1:
                return ("mkv", None)
            return ("mkv", p + ln + val)
        return ("mkv", None)
    if len(b) >= 12 and b[4:8] == b"ftyp":
        off = 0
        while off + 8 <= len(b):
            sz = struct.unpack_from(">I", b, off)[0]
            typ = b[off + 4:off + 8]
            if sz == 1:
                if off + 16 > len(b):
                    return ("mp4", None)
                sz = struct.unpack_from(">Q", b, off + 8)[0]
            if sz == 0:
                return ("mp4", None)
            if typ == b"mdat":
                return ("mp4", off + sz)
            off += sz
        return ("mp4", None)
    return ("unknown", None)


def _magic_type(head: bytes) -> str:
    for sig, ext in _MAGIC:
        if head.startswith(sig):
            return ext
    if len(head) >= 8 and head[4:8] == b"ftyp":
        return "mp4"
    return "bin"


# ── 목록 크롤 ──────────────────────────────────────────────────────────────
def fetch_total(menu: str, license_codes: str, provider: Optional[str]) -> Optional[int]:
    ep, mn = MENU[menu]
    p = {"menuNo": mn, "pageIndex": 1, "pageUnit": 12}
    if license_codes:
        p["licenseCd"] = license_codes
    if provider:
        p["searchSrcTrgetInttCd"] = provider
    try:
        r = sess().get(f"{BASE}/gongu/wrt/wrtCl/{ep}.do", params=p, timeout=30)
        m = _TOTAL.search(r.text)
        return int(m.group(1).replace(",", "")) if m else None
    except Exception:
        return None


def crawl_listing(menu: str, license_codes: str, limit: int, pages: List[int],
                  page_unit: int, provider: Optional[str], sleep: float,
                  max_pages_per_start: int = 400) -> List[str]:
    """지정된 시작페이지들에서 라운드로빈으로 wrtSn 수집(층화 샘플링)."""
    ep, mn = MENU[menu]
    url = f"{BASE}/gongu/wrt/wrtCl/{ep}.do"
    seen: Dict[str, None] = {}
    cursors = {p: p for p in pages}
    dead = set()
    while len(seen) < limit and len(dead) < len(pages):
        for start in pages:
            if len(seen) >= limit or start in dead:
                continue
            pg = cursors[start]
            if pg - start >= max_pages_per_start:
                dead.add(start); continue
            params = {"menuNo": mn, "pageIndex": pg, "pageUnit": page_unit}
            if license_codes:
                params["licenseCd"] = license_codes
            if provider:
                params["searchSrcTrgetInttCd"] = provider
            try:
                r = sess().get(url, params=params, timeout=30)
                r.raise_for_status()
                found = list(dict.fromkeys(_WRTSN.findall(r.text)))
            except Exception as e:
                print(f"  ⚠ listing p{pg} 실패: {e}"); found = []
            new = [w for w in found if w not in seen]
            if not found:
                dead.add(start)
            for w in new:
                if len(seen) < limit:
                    seen[w] = None
            cursors[start] = pg + 1
            time.sleep(sleep)
    return list(seen.keys())


def parse_detail(text: str) -> Dict[str, object]:
    fields: Dict[str, str] = {}
    for dt, dd in _DL.findall(text):
        label = _clean(dt)
        if label and label not in fields:
            fields[label] = dd

    def val(label: str) -> str:
        return _clean(fields.get(label, ""))

    lic_raw = fields.get("이용조건", "")
    m = _LICENSE_IMG.search(lic_raw)
    code = m.group(1) if m else ""
    author = re.sub(r"\s*\(저작물\s*[\d,]+\s*건\)\s*$", "", val("저작(권)자")).strip()
    wonmun = val("원문제공")
    hosted = "원문파일명" in wonmun
    external = ("원문URL" in wonmun) and not hosted
    return {
        "제목": val("저작물명"), "저작권자": author, "출처": val("출처"),
        "license_code": code, "license_name": LICENSE_NAME.get(code, ""),
        "요약정보": val("요약정보"),
        "분류_장르": val("분류(장르)") or val("분류"),
        "원문제공": wonmun, "hosted": hosted, "external": external,
        # external 인데 URL 값조차 비어있는지(= 완전 메타데이터-온리 레코드)
        "external_url_empty": external and wonmun.strip() in ("원문URL", "원문URL "),
        "has_dl_button": "wrtFileDownload" in text,
    }


def decode_filename(cd: str) -> str:
    """%-인코딩 우선(어문 전역·이미지 다수), 실패 시 latin-1→utf-8 mojibake 보정."""
    if not cd:
        return ""
    m = re.search(r"filename\*\s*=\s*UTF-8''([^;]+)", cd, re.I)
    if m:
        return unquote(m.group(1))
    m = re.search(r'filename\s*=\s*"?([^";]+)"?', cd)
    if not m:
        return ""
    raw = m.group(1).strip()
    if "%" in raw:                                   # P2: %-인코딩 먼저
        try:
            d = unquote(raw)
            if d and d != raw:
                return d
        except Exception:
            pass
    try:
        return raw.encode("latin-1").decode("utf-8")  # raw UTF-8 mojibake
    except (UnicodeDecodeError, UnicodeEncodeError):
        return raw


def _safe(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|\r\n\t]', "_", name)[:150]


def download_one_file(wrtsn: str, fsn: int, out_dir: Path, max_bytes: int,
                      chunk: int = 1 << 20) -> Dict[str, object]:
    """단일 fileSn 스트리밍 다운로드. 반환 dict 에 status/issue 포함."""
    url = f"{BASE}/gongu/wrt/cmmn/wrtFileDownload.do"
    res: Dict[str, object] = {"fileSn": fsn}
    part: Optional[Path] = None
    try:
        r = sess().get(url, params={"wrtSn": wrtsn, "fileSn": fsn},
                       timeout=(10, 300), stream=True)
        res["http_status"] = r.status_code
        if r.status_code != 200:
            res.update(status="failed", issue="http_error",
                       issue_detail=f"HTTP {r.status_code}")
            return res
        it = r.iter_content(chunk_size=chunk)
        head = b""
        for c in it:                                  # 최소 96B 확보 후 sentinel 검사
            head += c
            if len(head) >= _SENTINEL_MIN:
                break
        if not head:
            res.update(status="failed", issue="no_file_soft404",
                       issue_detail="empty body")
            return res
        if len(head) < 400 and _FNF.search(head.decode("latin-1", "ignore")):
            res.update(status="missing", issue="no_file_soft404",
                       issue_detail="File Not Found sentinel")
            return res
        fmt, declared = size_from_buf(head)
        res["declared_bytes"] = declared
        res["container"] = fmt
        if max_bytes and declared and declared > max_bytes:
            r.close()
            res.update(status="skipped", issue="size_exceeded",
                       issue_detail=f"{declared/2**20:.0f}MB > {max_bytes/2**20:.0f}MB cap")
            return res
        name = decode_filename(r.headers.get("Content-Disposition", "")) \
            or f"{wrtsn}_{fsn}.{_magic_type(head[:16])}"
        part = out_dir / f"{wrtsn}_{fsn}_{_safe(name)}.part"
        h = hashlib.sha1()
        total = 0
        with open(part, "wb") as f:
            f.write(head); h.update(head); total = len(head)
            for c in it:
                f.write(c); h.update(c); total += len(c)
                if max_bytes and total > max_bytes:
                    r.close()
                    f.close(); part.unlink(missing_ok=True)
                    res.update(status="skipped", issue="size_exceeded",
                               issue_detail=f">{max_bytes/2**20:.0f}MB while streaming")
                    return res
        final = part.with_suffix("")           # strip .part
        part.replace(final)
        res.update(status="ok", filename=name, path=str(final), size=total,
                   magic=_magic_type(head[:16]), sha1=h.hexdigest())
        if declared and total < declared * 0.98:
            res.update(status="ok", issue="truncated",
                       issue_detail=f"got {total} of ~{declared}")
        return res
    except requests.exceptions.Timeout as e:
        res.update(status="failed", issue="timeout", issue_detail=str(e)[:120])
    except requests.exceptions.ConnectionError as e:
        res.update(status="failed", issue="conn_error", issue_detail=str(e)[:120])
    except OSError as e:
        res.update(status="failed", issue="write_error", issue_detail=str(e)[:120])
    except Exception as e:
        res.update(status="failed", issue=type(e).__name__, issue_detail=str(e)[:120])
    finally:
        if part is not None and part.exists():
            part.unlink(missing_ok=True)
    return res


def download_work(wrtsn: str, out_dir: Path, max_bytes: int, max_files: int,
                  sleep: float) -> List[Dict[str, object]]:
    """fileSn=1..N 루프. sha1 중복(어문 hwp/txt byte-identical) 제거."""
    got: List[Dict[str, object]] = []
    seen_sha: set = set()
    for fsn in range(1, max_files + 1):
        r = download_one_file(wrtsn, fsn, out_dir, max_bytes)
        if r.get("status") == "missing":
            break                                   # 더 이상 파일 없음
        if r.get("status") == "ok":
            sha = r.get("sha1")
            if sha in seen_sha:                     # P10 동일내용 중복
                Path(str(r["path"])).unlink(missing_ok=True)
                r.update(status="skipped", issue="duplicate_sha1",
                         issue_detail="identical to earlier fileSn")
            else:
                seen_sha.add(sha)
        got.append(r)
        if r.get("status") == "failed":
            break
        time.sleep(sleep)
    return got


def existing_files(out_dir: Path, wrtsn: str) -> List[Path]:
    """재개용. .part(미완) 는 제외."""
    return [p for p in out_dir.glob(f"{wrtsn}_*")
            if p.suffix != ".part" and p.stat().st_size > 0]


# ── 기록 (증분 JSONL + 최종 xlsx) ──────────────────────────────────────────
class Recorder:
    def __init__(self, out: Path):
        self.out = out
        self.rec_path = out / "records.jsonl"
        self.iss_path = out / "issues.jsonl"
        self.lock = threading.Lock()
        self.records: List[Dict[str, object]] = []
        self.done_sn: set = set()
        # 재개: 기존 기록 로드. ★ 'failed' 는 done 으로 치지 않는다 —
        #   서버 503/네트워크 오류는 재시도 대상이어야 하므로.
        #   ok/skipped/dry_run(성공) + external/filtered(재시도 무의미) 만 종결 처리.
        TERMINAL = {"ok", "skipped_existing", "dry_run", "external", "filtered"}
        if self.rec_path.exists():
            for line in self.rec_path.read_text(encoding="utf-8").splitlines():
                try:
                    r = json.loads(line)
                    self.records.append(r)
                    if str(r.get("status")) in TERMINAL:
                        self.done_sn.add(str(r.get("wrtSn")))
                except Exception:
                    pass

    def add(self, rec: Dict[str, object]) -> None:
        with self.lock:
            self.records.append(rec)
            with open(self.rec_path, "a", encoding="utf-8") as f:
                f.write(json.dumps(rec, ensure_ascii=False) + "\n")
            if rec.get("issue"):
                with open(self.iss_path, "a", encoding="utf-8") as f:
                    f.write(json.dumps({k: rec.get(k) for k in
                                        ("wrtSn", "menu", "license_code", "제목",
                                         "status", "issue", "issue_detail",
                                         "http_status", "원문제공", "detail_url")},
                                       ensure_ascii=False) + "\n")

    def write_xlsx(self, meta: Dict[str, object]) -> None:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "gongu_index"
        cols = ["wrtSn", "menu", "제목", "저작권자", "출처", "license_code", "license_name",
                "요약정보", "desc_len", "분류_장르", "원문제공", "hosted",
                "file_count", "file_names", "saved_paths", "total_bytes",
                "status", "issue", "issue_detail", "http_status", "detail_url"]
        ws.append(cols)
        # wrtSn upsert (마지막 기록 우선)
        merged: Dict[str, Dict] = {}
        for r in self.records:
            merged[str(r.get("wrtSn"))] = r
        for r in merged.values():
            ws.append([r.get(c, "") for c in cols])
        wi = wb.create_sheet("issues")
        wi.append(["wrtSn", "제목", "status", "issue", "issue_detail", "http_status",
                   "원문제공", "detail_url"])
        for r in merged.values():
            if r.get("issue"):
                wi.append([r.get("wrtSn"), r.get("제목"), r.get("status"), r.get("issue"),
                           r.get("issue_detail"), r.get("http_status"),
                           r.get("원문제공"), r.get("detail_url")])
        wsum = wb.create_sheet("summary")
        for k, v in meta.items():
            wsum.append([k, str(v)])
        c = Counter(str(r.get("status")) for r in merged.values())
        wsum.append([]); wsum.append(["status", "count"])
        for k, v in c.most_common():
            wsum.append([k, v])
        ci = Counter(str(r.get("issue")) for r in merged.values() if r.get("issue"))
        wsum.append([]); wsum.append(["issue", "count"])
        for k, v in ci.most_common():
            wsum.append([k, v])
        wb.save(self.out / "gongu_index.xlsx")


def main() -> int:
    ap = argparse.ArgumentParser(description="공유마당 저작물 다운로더(익명 HTTP)")
    ap.add_argument("--menu", choices=list(MENU), default="image")
    ap.add_argument("--license", default="", help="licenseCd (만료 97 / 기증 98,99 / CCL 20..27 / KOGL 01,02,03,04)")
    ap.add_argument("--provider", default=None, help="searchSrcTrgetInttCd")
    ap.add_argument("--limit", type=int, default=30)
    ap.add_argument("--pages", default="1", help="시작 페이지 목록(쉼표) — 층화 샘플링. 예: 1,700,1400")
    ap.add_argument("--page-unit", type=int, default=100)
    ap.add_argument("--wrtsn-file", default="", help="사전 선정된 wrtSn 목록 파일(줄바꿈 구분) — 목록크롤 생략")
    ap.add_argument("--out", default="dataset/gongu")
    ap.add_argument("--require-desc", action="store_true")
    ap.add_argument("--desc-min", type=int, default=20)
    ap.add_argument("--exclude-title", default="", help="제목 정규식 매치 시 제외(기본 미적용)")
    ap.add_argument("--hosted-only", action="store_true", help="external(원문URL)은 다운로드 생략하고 사유 기록")
    ap.add_argument("--max-file-mb", type=int, default=0, help="파일당 상한(MB). 0=무제한")
    ap.add_argument("--max-files", type=int, default=12, help="작품당 fileSn 최대 탐색")
    ap.add_argument("--workers", type=int, default=4, help="동시 처리 수(최대 8)")
    ap.add_argument("--dry-run", action="store_true", help="다운로드 없이 메타데이터만")
    ap.add_argument("--sleep", type=float, default=0.5)
    ap.add_argument("--retries", type=int, default=3)
    ap.add_argument("--no-abort", action="store_true", help="external-only 자동중단 비활성")
    args = ap.parse_args()

    if args.workers > 8:
        print("⚠ workers 8 초과 금지 — 8로 제한"); args.workers = 8
    _, menu_no = MENU[args.menu]
    out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
    max_bytes = args.max_file_mb * 1024 * 1024
    rec = Recorder(out)
    t_start = time.time()

    # ── 대상 wrtSn 확보 ──
    if args.wrtsn_file:
        wrtsns = [l.strip() for l in Path(args.wrtsn_file).read_text().splitlines()
                  if l.strip().isdigit()][: args.limit]
        total_avail = len(wrtsns)
        print(f"[1/2] wrtSn 파일 로드: {len(wrtsns)}건 ({args.wrtsn_file})")
    else:
        pages = [int(p) for p in args.pages.split(",") if p.strip().isdigit()] or [1]
        total_avail = fetch_total(args.menu, args.license, args.provider)
        over = args.limit * 3 if (args.hosted_only or args.require_desc) else args.limit
        print(f"[1/2] 목록 크롤 menu={args.menu} license={args.license or '전체'} "
              f"총 {total_avail if total_avail is not None else '?'}건 | 시작페이지 {pages} "
              f"pageUnit={args.page_unit} → 최대 {over}건 수집")
        wrtsns = crawl_listing(args.menu, args.license, over, pages,
                               args.page_unit, args.provider, args.sleep)
        print(f"      wrtSn {len(wrtsns)}건 수집(필터 전)")

    todo = [w for w in wrtsns if w not in rec.done_sn]
    if len(todo) < len(wrtsns):
        print(f"      재개: 기존 {len(wrtsns)-len(todo)}건 건너뜀")

    # ── 처리 ──
    print(f"[2/2] 상세 파싱 + {'스캔' if args.dry_run else '다운로드'} "
          f"(목표 {args.limit}건, workers={args.workers}"
          f"{f', 상한 {args.max_file_mb}MB' if args.max_file_mb else ''}) ...")
    # 재개 시 이미 확보한 건수를 limit 에 반영(재실행마다 limit 만큼 더 받지 않도록)
    _prior_ok = sum(1 for r in rec.records
                    if str(r.get("status")) in ("ok", "skipped_existing", "dry_run"))
    if _prior_ok:
        print(f"      기존 확보 {_prior_ok}건을 목표 {args.limit}건에 산입")
    state = {"accepted": _prior_ok, "external": 0, "checked": 0, "bytes": 0, "n": 0}
    lock = threading.Lock()
    stop = threading.Event()
    title_re = re.compile(args.exclude_title) if args.exclude_title else None

    def handle(wrtsn: str) -> None:
        if stop.is_set():
            return
        with lock:
            if state["accepted"] >= args.limit:
                stop.set(); return
        durl = f"{BASE}/gongu/wrt/wrt/view.do?wrtSn={wrtsn}&menuNo={menu_no}"
        meta = None
        for attempt in range(args.retries):
            try:
                r = sess().get(f"{BASE}/gongu/wrt/wrt/view.do",
                               params={"wrtSn": wrtsn, "menuNo": menu_no}, timeout=30)
                if r.status_code in (429, 503):      # 서버 과부하 — 지수 백오프
                    err = f"HTTP {r.status_code} (server busy)"
                    time.sleep(min(60, 5 * (2 ** attempt)))
                    continue
                r.raise_for_status()
                meta = parse_detail(r.text); break
            except Exception as e:
                err = str(e)[:100]
                time.sleep(min(30, 2 * (2 ** attempt)))
        if meta is None:
            rec.add({"wrtSn": wrtsn, "menu": args.menu, "status": "failed",
                     "issue": "detail_fail", "issue_detail": err, "detail_url": durl})
            return
        desc = str(meta["요약정보"])
        base = {"wrtSn": wrtsn, "menu": args.menu, **meta, "desc_len": len(desc),
                "detail_url": durl, "file_count": 0, "file_names": "",
                "saved_paths": "", "total_bytes": 0}

        with lock:
            state["checked"] += 1
            if meta["external"]:
                state["external"] += 1
            # P7: external-only 셀 자동 중단 (메시지는 1회만)
            if (not args.no_abort and state["checked"] >= 20
                    and state["external"] / state["checked"] > 0.95
                    and not stop.is_set()):
                stop.set()
                print(f"\n🛑 EXTERNAL-ONLY 셀로 판정 — {state['external']}/{state['checked']} "
                      f"가 원문URL(파일 없음). 중단하고 사유를 기록합니다.")

        if title_re and title_re.search(str(meta["제목"])):
            rec.add({**base, "status": "filtered", "issue": "excluded_title"}); return
        if args.require_desc and len(desc) < args.desc_min:
            rec.add({**base, "status": "filtered", "issue": "desc_too_short",
                     "issue_detail": f"{len(desc)} < {args.desc_min}"}); return
        if meta["external"]:
            detail = ("원문URL 값 비어있음(메타데이터-온리)" if meta.get("external_url_empty")
                      else "외부 기관 호스팅")
            rec.add({**base, "status": "external", "issue": "external_no_file",
                     "issue_detail": detail}); return
        if not meta["hosted"] and args.hosted_only:
            rec.add({**base, "status": "external", "issue": "no_hosted_marker",
                     "issue_detail": f"원문제공='{meta['원문제공'][:40]}'"}); return

        if args.dry_run:
            with lock: state["accepted"] += 1
            rec.add({**base, "status": "dry_run"}); return

        ex = existing_files(out, wrtsn)
        if ex:
            with lock: state["accepted"] += 1
            rec.add({**base, "status": "skipped_existing", "file_count": len(ex),
                     "file_names": ";".join(p.name for p in ex),
                     "saved_paths": ";".join(str(p) for p in ex),
                     "total_bytes": sum(p.stat().st_size for p in ex)}); return

        with lock:  # 다운로드 직전 재확인 — 오버슈트 최소화(최대 workers-1)
            if state["accepted"] >= args.limit:
                stop.set(); return
        files = download_work(wrtsn, out, max_bytes, args.max_files, min(0.3, args.sleep))
        good = [f for f in files if f.get("status") == "ok"]
        issues = [f for f in files if f.get("issue")]
        nbytes = sum(int(f.get("size") or 0) for f in good)
        r = {**base, "file_count": len(good),
             "file_names": ";".join(str(f.get("filename")) for f in good),
             "saved_paths": ";".join(str(f.get("path")) for f in good),
             "total_bytes": nbytes}
        if good:
            r["status"] = "ok"
            with lock:
                state["accepted"] += 1; state["bytes"] += nbytes
        else:
            r["status"] = "failed"
        if issues:
            r["issue"] = issues[0].get("issue")
            r["issue_detail"] = issues[0].get("issue_detail")
            r["http_status"] = issues[0].get("http_status")
        rec.add(r)

        with lock:
            state["n"] += 1
            if state["n"] % 10 == 0 or state["accepted"] >= args.limit:
                el = time.time() - t_start
                rate = state["accepted"] / el if el else 0
                eta = (args.limit - state["accepted"]) / rate / 60 if rate else 0
                print(f"  [{state['accepted']}/{args.limit}] "
                      f"{state['bytes']/2**30:.2f}GB | ext {state['external']} | "
                      f"{el/60:.0f}m 경과 · ETA {eta:.0f}m")

    from concurrent.futures import ThreadPoolExecutor
    try:
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            list(pool.map(handle, todo))
    except KeyboardInterrupt:
        print("\n⏸ 중단됨 — records.jsonl 로 재개 가능")

    meta = {"menu": args.menu, "license": args.license, "limit": args.limit,
            "pages": args.pages, "page_unit": args.page_unit,
            "total_available": total_avail, "workers": args.workers,
            "max_file_mb": args.max_file_mb, "out": str(out),
            "elapsed_min": round((time.time() - t_start) / 60, 1),
            "downloaded_GB": round(state["bytes"] / 2**30, 2),
            "hosted_rate": (f"{100*(state['checked']-state['external'])/state['checked']:.1f}%"
                            if state["checked"] else "n/a")}
    rec.write_xlsx(meta)
    merged = {str(r.get("wrtSn")): r for r in rec.records}
    c = Counter(str(r.get("status")) for r in merged.values())
    ci = Counter(str(r.get("issue")) for r in merged.values() if r.get("issue"))
    print(f"\n=== 요약 === {dict(c)}")
    if ci:
        print(f"  이슈: {dict(ci)}")
    print(f"  hosted율 {meta['hosted_rate']} | 수집 {state['bytes']/2**30:.2f}GB | "
          f"{meta['elapsed_min']}분")
    print(f"  인덱스 {out}/gongu_index.xlsx (+ issues 시트) | 증분 {out}/records.jsonl")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

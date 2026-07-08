"""
KOGL 데이터셋 빌더 — SELECT → DOWNLOAD → MANIFEST.

붙임1 원문저작물 메타데이터 Excel을 read-only로 읽어, 저작권자명이 채워진 행을
분류(이미지/영상/어문)별로 재현 가능하게 샘플링하고, 공개 썸네일을 내려받아
Pillow로 디코딩 검증한 뒤, 계약서 생성용 매니페스트(manifest.xlsx)를 만든다.

공개 썸네일 URL(검증된 패턴):
    https://www.kogl.or.kr  +  urllib.parse.quote(쎔네일웹경로, safe="/")
원본 파일은 로그인이 필요해 내려받지 않는다. 썸네일(330×230)만 받는다.

Usage:
    # 기본: 이미지 400 / 영상 300 / 어문 300, dataset/ 에 출력
    python -m api.module.dataset_builder.build

    # 카운트 변경
    python -m api.module.dataset_builder.build --counts 이미지=400,영상=300,어문=300

    # 드라이런(유형별 N개만) — 가벼운 실 다운로드 테스트
    python -m api.module.dataset_builder.build --dry-run 1

    # SELECT 단계만 (다운로드/매니페스트 생략)
    python -m api.module.dataset_builder.build --select-only

    # 동시 다운로드 워커 수 / 시드 / 출력 경로
    python -m api.module.dataset_builder.build --workers 5 --seed 42 --out dataset

Outputs:
    {out}/selection.csv                          (선택된 행의 모든 메타데이터)
    {out}/images|documents|videos/{원문인덱스}.{jpg,png}   (썸네일)
    {out}/manifest.xlsx                          (메타데이터 + 다운로드 상태)
"""

from __future__ import annotations

import argparse
import csv
import random
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
from pathlib import Path

import requests
import urllib.parse
import urllib3
from PIL import Image

# 프로젝트 루트를 import 경로에 추가 (스크립트로 직접 실행될 때) ----------------
sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

THIS_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_XLSX = PROJECT_ROOT / "docs" / "붙임1_원문저작물 메타데이터.xlsx"

KOGL_BASE = "https://www.kogl.or.kr"
USER_AGENT = "soongsil-research-crawler/1.0 (academic; copyright-metadata)"
MAX_MEDIA_BYTES = 50 * 1024 * 1024  # 50MB — 직접 미디어 URL 방어 가드

# 기본 타깃 카운트 (분류 → 개수) -------------------------------------------------
DEFAULT_COUNTS = {"이미지": 400, "영상": 300, "어문": 300}

# 분류 → 저장 하위폴더 ----------------------------------------------------------
SUBFOLDER = {"이미지": "images", "어문": "documents", "영상": "videos"}

# 매니페스트/셀렉션에 보존할 메타데이터 컬럼 (오타 주의: '쎔'네일웹경로) -----------
METADATA_COLUMNS = [
    "원문인덱스",
    "분류",
    "정보유형",
    "제목",
    "쎔네일웹경로",       # 오타가 원본 그대로임 (썸 아님)
    "게시글URL",
    "원본파일명",
    "동영상URL주소",
    "저작권자명",
    "저작권자 소속",
    "공동저작자",
    "저작인접권자",
    "저작물성격",
    "비보호저작물",
    "업무상저작물",
    "상업적이용허락",
    "복제권",
    "저작재산권 공연권",
    "공중송신권",
    "저작재산권 전시권",
    "저작재산권 배포권",
    "대여권",
    "2차적저작물 작성권",
    "공공누리 유형",
    "언어",
    "제작일자",
    "공표일자",
    "계약상 유효기간",
    "저작권 만료일",
    "초상권",
    "주제어",
    "해시태그",
    "원본소유자",
]

# 완성도 점수에 쓰는 선택적 계약 필드 (많이 채워질수록 계약서 생성에 유리) --------
COMPLETENESS_FIELDS = [
    "계약상 유효기간",
    "공동저작자",
    "저작인접권자",
    "저작권 만료일",
    "제작일자",
    "공표일자",
    "주제어",
    "해시태그",
]

# 다운로드 상태에 따라 매니페스트에 추가되는 컬럼 -------------------------------
EXTRA_COLUMNS = ["local_path", "download_status", "media_available", "subfolder"]

# 매니페스트 출력에서 제외할 컬럼 (내부 다운로드용으로만 필요)
MANIFEST_EXCLUDE = {"쎔네일웹경로"}

# KOGL 인증서 검증을 끄므로 경고를 한 번만 억제한다.
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ---------------------------------------------------------------- helpers ---
def clean(value: object) -> str:
    """셀 값을 문자열로 정규화하고, 작은따옴표로 감싼 경우 한 겹 벗긴다."""
    if value is None:
        return ""
    s = str(value).strip()
    if len(s) >= 2 and s[0] == "'" and s[-1] == "'":
        s = s[1:-1].strip()
    return s


def build_thumbnail_url(thumb_path: str) -> str:
    """쎔네일웹경로 -> 공개 썸네일 URL (검증된 패턴)."""
    return KOGL_BASE + urllib.parse.quote(thumb_path, safe="/")


def thumb_ext(thumb_path: str) -> str:
    """쎔네일웹경로 확장자(.jpg/.png)를 소문자로 반환. 없으면 .jpg."""
    ext = Path(thumb_path).suffix.lower()
    return ext if ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp") else ".jpg"


def completeness_score(rec: dict[str, str]) -> int:
    """선택적 계약 필드가 채워진 개수 — 높을수록 계약서 생성에 풍부한 행."""
    return sum(1 for f in COMPLETENESS_FIELDS if rec.get(f))


def decodes_ok(path: Path) -> bool:
    """이미 저장된 파일이 Pillow로 디코딩되는지 검사 (resumable 판정용)."""
    try:
        with Image.open(path) as im:
            im.load()
        return True
    except Exception:  # noqa: BLE001
        return False


# ---------------------------------------------------------------- loading ---
# 저작권자명 placeholder 값 (계약서 생성에 쓸 수 없는 값) — 선택에서 제외
_PLACEHOLDER_OWNERS = {"-", "--", ".", "미상", "없음", "N/A", "n/a", "해당없음"}


def read_candidates(xlsx_path: Path, target_classes: set[str]):
    """
    워크북을 read-only로 스트리밍하며 분류별 후보를 모은다.

    저작권자명이 비어있지 않은 행만(계약용 실권리 데이터) 유지하고,
    타깃 분류에 속하는 행만 모은다.

    반환: (groups, stats)
      groups: dict[분류 -> list[dict(METADATA_COLUMNS -> cleaned str)]]
      stats:  dict (총 행수, 권리없음 제외, 후보 등)
    """
    import openpyxl  # 지연 import (무거운 의존성)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = [clean(h) for h in next(rows_iter)]

    # 필요한 컬럼 인덱스 매핑 (없으면 즉시 오류)
    col_idx: dict[str, int] = {}
    missing: list[str] = []
    for name in METADATA_COLUMNS:
        if name in header:
            col_idx[name] = header.index(name)
        else:
            missing.append(name)
    if missing:
        wb.close()
        raise KeyError(
            f"Excel 헤더에서 다음 컬럼을 찾지 못했습니다: {missing}"
        )

    groups: dict[str, list[dict[str, str]]] = {c: [] for c in target_classes}
    stats = {"total_rows": 0, "skipped_no_rights": 0, "kept": 0, "no_thumb_in_kept": 0}

    for raw in rows_iter:
        stats["total_rows"] += 1
        cls = clean(raw[col_idx["분류"]])
        if cls not in target_classes:
            continue
        owner = clean(raw[col_idx["저작권자명"]])
        # 저작권자명이 비었거나 placeholder('-' 등)면 제외 — 계약서에 실제 권리자명이
        # 필요하므로 (이미지의 ~27%가 '-' placeholder. 실명 풀이 목표치보다 훨씬 큼)
        if not owner or owner in _PLACEHOLDER_OWNERS:
            stats["skipped_no_rights"] += 1
            continue
        rec = {name: clean(raw[col_idx[name]]) for name in METADATA_COLUMNS}
        # 공공누리 코드 9 → 0 정규화: KOGL이 '제0유형'(공개 라이선스 미적용/별도
        # 이용허락 대상)으로 표시하므로 의미상 0이 맞다. 원본 붙임1 Excel은 9 그대로 보존됨.
        if rec.get("공공누리 유형") == "9":
            rec["공공누리 유형"] = "0"
        groups[cls].append(rec)
        stats["kept"] += 1
        if not rec["쎔네일웹경로"]:
            stats["no_thumb_in_kept"] += 1

    wb.close()
    return groups, stats


# ----------------------------------------------------------------- select ---
def select_for_class(candidates: list[dict[str, str]], count: int, seed: int):
    """
    한 분류 후보에서 count개를 재현 가능하게 선택한다.

    전략: 시드로 한 번 섞은 뒤(무작위성) 완성도 점수 내림차순으로 안정 정렬해
    상위 "더 풍부한 부분집합"(count*3 또는 전체)을 만들고, 거기서 시드 무작위
    추출한다. 즉 풍부한 행을 선호하되 약간의 무작위성을 유지한다.
    풀이 count보다 작으면 전부 반환한다.
    """
    rng = random.Random(seed)
    shuffled = candidates[:]
    rng.shuffle(shuffled)
    # 안정 정렬 → 동점은 섞인 순서를 유지하므로 시드 재현성이 보장된다.
    shuffled.sort(key=completeness_score, reverse=True)

    if len(shuffled) <= count:
        return list(shuffled)

    richer_size = min(len(shuffled), max(count * 3, count))
    richer = shuffled[:richer_size]
    return rng.sample(richer, count)


def write_selection_csv(out_path: Path, selected_by_class: dict[str, list[dict[str, str]]]):
    """선택 행을 모든 메타데이터 컬럼과 함께 selection.csv(utf-8-sig)로 기록."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    n = 0
    with out_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=METADATA_COLUMNS)
        writer.writeheader()
        for cls in sorted(selected_by_class):
            for rec in selected_by_class[cls]:
                writer.writerow({k: rec.get(k, "") for k in METADATA_COLUMNS})
                n += 1
    return n


# --------------------------------------------------------------- download ---
def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    s.verify = False  # KOGL 인증서 검증 무시 (build_eval_manifest와 동일)
    return s


def _content_length_too_big(session: requests.Session, url: str) -> bool:
    """
    직접 미디어 URL용 50MB 방어 가드: HEAD로 Content-Length를 확인.
    썸네일은 매우 작아 절대 트리거되지 않는다(이 가드는 향후 직접 미디어용).
    """
    try:
        r = session.head(url, timeout=15, allow_redirects=True)
        cl = r.headers.get("Content-Length")
        if cl and int(cl) > MAX_MEDIA_BYTES:
            return True
    except Exception:  # noqa: BLE001
        pass  # HEAD 실패 시엔 GET 단계에서 처리
    return False


def _is_transient(exc: Exception | None, status: int | None) -> bool:
    """일시적 오류(재시도 대상)인지 판정: 연결/타임아웃/5xx/429."""
    if exc is not None:
        return isinstance(exc, (requests.ConnectionError, requests.Timeout))
    if status is not None:
        return status == 429 or 500 <= status < 600
    return False


def download_one(
    session: requests.Session,
    rec: dict[str, str],
    out_dir: Path,
    rng: random.Random,
    is_thumbnail: bool = True,
) -> dict[str, str]:
    """
    한 작품의 썸네일을 내려받아 검증한다.

    반환 dict: download_status / local_path(상대, 또는 "") / media_available
      - skipped_existing : 이미 존재하고 디코딩되어 재다운로드 생략
      - ok               : 새로 받고 디코딩 검증 통과
      - decode_error     : 받았으나 디코딩 실패(파일 삭제)
      - failed           : 썸네일 경로 없음/HTTP 오류/네트워크 오류
      - skipped_large    : 50MB 초과 직접 미디어(메타데이터만, 썸네일엔 비발생)
    """
    idx = rec["원문인덱스"]
    thumb_path = rec["쎔네일웹경로"]
    result = {"download_status": "failed", "local_path": "", "media_available": "False"}

    if not thumb_path:
        return result  # 썸네일 경로 자체가 없음

    cls = rec["분류"]
    sub = SUBFOLDER.get(cls, "misc")
    dest_dir = out_dir / sub
    dest_dir.mkdir(parents=True, exist_ok=True)
    ext = thumb_ext(thumb_path)
    dest = dest_dir / f"{idx}{ext}"
    rel = f"{sub}/{dest.name}"

    # resumable: 이미 존재하고 디코딩되면 건너뛴다
    if dest.exists() and decodes_ok(dest):
        result.update(download_status="skipped_existing", local_path=rel,
                      media_available="True")
        return result

    url = build_thumbnail_url(thumb_path)

    # 직접 미디어 URL 방어 가드(썸네일엔 비발생) — 50MB 초과 시 메타데이터만
    if not is_thumbnail and _content_length_too_big(session, url):
        result.update(download_status="skipped_large")
        return result

    # 정중함: 요청마다 0.2~0.6s 지터
    time.sleep(rng.uniform(0.2, 0.6))

    # 최대 3회, 지수 백오프로 재시도(일시적 오류만)
    last_err = ""
    for attempt in range(3):
        try:
            r = session.get(url, timeout=30)
            if r.status_code == 200:
                data = r.content
                # 디코딩 검증 (메모리에서 먼저 시도)
                try:
                    with Image.open(BytesIO(data)) as im:
                        im.load()
                except Exception:  # noqa: BLE001
                    # 파일로 저장하지 않고 decode_error 처리
                    if dest.exists():
                        dest.unlink(missing_ok=True)
                    result.update(download_status="decode_error")
                    return result
                dest.write_bytes(data)
                result.update(download_status="ok", local_path=rel,
                              media_available="True")
                return result
            # 비-200
            if _is_transient(None, r.status_code) and attempt < 2:
                time.sleep((2 ** attempt) * 0.5 + rng.uniform(0, 0.3))
                last_err = f"HTTP {r.status_code}"
                continue
            last_err = f"HTTP {r.status_code}"
            break
        except Exception as e:  # noqa: BLE001
            last_err = f"{type(e).__name__}"
            if _is_transient(e, None) and attempt < 2:
                time.sleep((2 ** attempt) * 0.5 + rng.uniform(0, 0.3))
                continue
            break

    result["_error"] = last_err
    return result


def download_all(
    selected_by_class: dict[str, list[dict[str, str]]],
    out_dir: Path,
    workers: int,
    seed: int,
):
    """모든 선택 행의 썸네일을 동시 다운로드. 결과를 원문인덱스 기준 dict로 반환."""
    tasks: list[dict[str, str]] = []
    for cls in sorted(selected_by_class):
        tasks.extend(selected_by_class[cls])

    results: dict[str, dict[str, str]] = {}
    session = _session()
    # 워커별 독립 RNG(시드+오프셋) — 지터/백오프 재현성 유지
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {}
        for i, rec in enumerate(tasks):
            rng = random.Random(seed + i)
            futures[ex.submit(download_one, session, rec, out_dir, rng)] = rec
        done = 0
        total = len(futures)
        for fut in as_completed(futures):
            rec = futures[fut]
            try:
                res = fut.result()
            except Exception as e:  # noqa: BLE001
                res = {"download_status": "failed", "local_path": "",
                       "media_available": "False", "_error": type(e).__name__}
            results[rec["원문인덱스"]] = res
            done += 1
            if done % 50 == 0 or done == total:
                print(f"  다운로드 진행 {done}/{total}")
    session.close()
    return results


# --------------------------------------------------------------- manifest ---
def write_manifest(
    out_path: Path,
    selected_by_class: dict[str, list[dict[str, str]]],
    download_results: dict[str, dict[str, str]],
):
    """manifest.xlsx 작성: 메타데이터 컬럼 + local_path/download_status/media_available/subfolder."""
    import openpyxl  # 지연 import

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "manifest"
    # 쎔네일웹경로는 다운로드 URL 생성에만 내부적으로 쓰이며 매니페스트 출력에는 불필요 → 제외
    manifest_meta_cols = [c for c in METADATA_COLUMNS if c not in MANIFEST_EXCLUDE]
    columns = manifest_meta_cols + EXTRA_COLUMNS
    ws.append(columns)

    n = 0
    for cls in sorted(selected_by_class):
        for rec in selected_by_class[cls]:
            idx = rec["원문인덱스"]
            dl = download_results.get(idx, {})
            row = [rec.get(k, "") for k in manifest_meta_cols]
            row.append(dl.get("local_path", ""))
            row.append(dl.get("download_status", ""))
            row.append(dl.get("media_available", "False"))
            row.append(SUBFOLDER.get(cls, "misc"))
            ws.append(row)
            n += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return n


def print_summary(
    counts: dict[str, int],
    selected_by_class: dict[str, list[dict[str, str]]],
    download_results: dict[str, dict[str, str]],
):
    """분류별 요청/ok/failed/decode_error/skipped 요약을 출력한다."""
    print()
    print("=== 다운로드 요약 (분류별) ===")
    print(f"{'분류':<6} {'요청':>6} {'ok':>6} {'skip':>6} {'decode':>7} {'failed':>7}")
    for cls in sorted(selected_by_class):
        recs = selected_by_class[cls]
        tally = {"ok": 0, "skipped_existing": 0, "decode_error": 0,
                 "failed": 0, "skipped_large": 0}
        for rec in recs:
            st = download_results.get(rec["원문인덱스"], {}).get("download_status", "failed")
            tally[st] = tally.get(st, 0) + 1
        print(f"{cls:<6} {counts.get(cls, 0):>6} {tally['ok']:>6} "
              f"{tally['skipped_existing']:>6} {tally['decode_error']:>7} "
              f"{tally['failed'] + tally['skipped_large']:>7}")


# ----------------------------------------------------------------- args -----
def parse_counts(spec: str) -> dict[str, int]:
    """'이미지=400,영상=300,어문=300' -> {'이미지':400,...}."""
    counts: dict[str, int] = {}
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "=" not in part:
            raise argparse.ArgumentTypeError(f"잘못된 카운트 형식: {part!r} (분류=개수)")
        cls, _, num = part.partition("=")
        cls = cls.strip()
        try:
            counts[cls] = int(num.strip())
        except ValueError:
            raise argparse.ArgumentTypeError(f"개수가 정수가 아님: {part!r}")
    return counts


# ------------------------------------------------------------------ main ----
def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="KOGL 데이터셋 빌더 (SELECT → DOWNLOAD → MANIFEST)."
    )
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX),
                        help="KOGL 메타데이터 Excel 경로")
    parser.add_argument(
        "--counts", type=parse_counts, default=dict(DEFAULT_COUNTS),
        help="분류별 개수 (기본 이미지=400,영상=300,어문=300)")
    parser.add_argument("--dry-run", type=int, default=0, metavar="N",
                        help="모든 분류를 N개로 덮어써 가벼운 실 다운로드 테스트")
    parser.add_argument("--out", default="dataset",
                        help="출력 디렉터리 (프로젝트 루트 기준, 기본 dataset)")
    parser.add_argument("--workers", type=int, default=5,
                        help="동시 다운로드 워커 수 (기본 5)")
    parser.add_argument("--seed", type=int, default=42,
                        help="샘플링/지터 시드 (재현성, 기본 42)")
    parser.add_argument("--select-only", action="store_true",
                        help="SELECT 단계만 실행(다운로드/매니페스트 생략)")
    args = parser.parse_args(argv)

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"입력 Excel을 찾을 수 없습니다: {xlsx_path}")
        return 1

    counts = dict(args.counts)
    if args.dry_run and args.dry_run > 0:
        counts = {cls: args.dry_run for cls in counts}
        print(f"[DRY-RUN] 모든 분류를 {args.dry_run}개로 제한합니다.")

    target_classes = set(counts)
    # 출력 경로: 절대경로면 그대로, 아니면 프로젝트 루트 기준
    out_dir = Path(args.out)
    if not out_dir.is_absolute():
        out_dir = PROJECT_ROOT / out_dir

    # --- Stage 1: SELECT ---
    print(f"Excel 읽는 중 (read-only): {xlsx_path}")
    print(f"타깃 분류/카운트: {counts}, seed={args.seed}")
    try:
        groups, stats = read_candidates(xlsx_path, target_classes)
    except KeyError as e:
        print(f"컬럼 오류: {e}")
        return 2

    print(
        f"스캔 완료: 총 {stats['total_rows']}행, "
        f"권리없음 {stats['skipped_no_rights']} 제외, "
        f"후보 {stats['kept']}행 (썸네일없는 후보 {stats['no_thumb_in_kept']})"
    )

    selected_by_class: dict[str, list[dict[str, str]]] = {}
    print("분류별 후보 / 선택:")
    for cls in counts:
        pool = groups.get(cls, [])
        want = counts[cls]
        if len(pool) < want:
            print(f"  ! {cls}: 후보 {len(pool)} < 요청 {want} — 전부 선택")
        sel = select_for_class(pool, want, args.seed)
        selected_by_class[cls] = sel
        print(f"  {cls}: 후보 {len(pool)} → 선택 {len(sel)}")

    sel_csv = out_dir / "selection.csv"
    n_sel = write_selection_csv(sel_csv, selected_by_class)
    print(f"selection.csv 작성: {sel_csv} ({n_sel}행)")

    if args.select_only:
        print("--select-only: 다운로드/매니페스트를 생략합니다.")
        return 0

    # --- Stage 2: DOWNLOAD ---
    print()
    print(f"썸네일 다운로드 시작 (workers={args.workers}) -> {out_dir}")
    download_results = download_all(selected_by_class, out_dir, args.workers, args.seed)

    # --- Stage 3: MANIFEST ---
    manifest_path = out_dir / "manifest.xlsx"
    n_manifest = write_manifest(manifest_path, selected_by_class, download_results)
    print(f"manifest.xlsx 작성: {manifest_path} ({n_manifest}행)")

    print_summary(counts, selected_by_class, download_results)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

"""
매니페스트 증분 확장: 기존 1000작품(400/300/300)을 보존하고 분류별 500개로 늘린다
(이미지 +100, 어문 +200, 영상 +200). 기존 idx 는 후보에서 제외해 중복 없이 신규만 추가.
신규 작품의 썸네일을 내려받고, 결합 매니페스트(manifest.xlsx)를 다시 쓴다. 원본은 별도
(download_originals_parallel)로 받는다. 기존 manifest 는 manifest_1000_backup.xlsx 로 백업.

사용: python -m api.module.dataset_builder.expand_manifest [--counts 이미지=500,어문=500,영상=500]
"""
from __future__ import annotations
import argparse, shutil
from collections import Counter
from pathlib import Path

from .build import (read_candidates, select_for_class, download_all,
                    SUBFOLDER, METADATA_COLUMNS, MANIFEST_EXCLUDE, EXTRA_COLUMNS,
                    DEFAULT_XLSX, PROJECT_ROOT, parse_counts)

TARGET_DEFAULT = {"이미지": 500, "어문": 500, "영상": 500}


def load_existing(manifest_path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(manifest_path); ws = wb.active
    hdr = [c.value for c in ws[1]]
    rows = [dict(zip(hdr, [c.value for c in r])) for r in ws.iter_rows(min_row=2)]
    wb.close()
    return hdr, rows


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="매니페스트 증분 확장 (분류별 목표치까지 신규 추가)")
    ap.add_argument("--counts", type=parse_counts, default=dict(TARGET_DEFAULT))
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--workers", type=int, default=5)
    ap.add_argument("--out", default=str(PROJECT_ROOT / "dataset"))
    args = ap.parse_args(argv)
    target = dict(args.counts)
    out_dir = Path(args.out)
    manifest_path = out_dir / "manifest.xlsx"

    hdr, existing = load_existing(manifest_path)
    existing_idx = {str(r["원문인덱스"]) for r in existing}
    by_cls = Counter(r["분류"] for r in existing)
    print(f"기존: {dict(by_cls)} (총 {len(existing)})")
    need = {c: max(0, target[c] - by_cls.get(c, 0)) for c in target}
    print(f"목표 {target} → 신규 필요 {need}")
    if sum(need.values()) == 0:
        print("추가할 작품 없음 — 종료"); return 0

    groups, stats = read_candidates(Path(DEFAULT_XLSX), set(target))
    print(f"144k 스캔: 총 {stats['total_rows']}행, 권리없음 {stats['skipped_no_rights']} 제외, 후보 {stats['kept']}")
    new_sel = {}
    for cls in target:
        pool = [c for c in groups.get(cls, []) if c["원문인덱스"] not in existing_idx]
        if need[cls] and len(pool) < need[cls]:
            print(f"  ! {cls}: 신규후보 {len(pool)} < 필요 {need[cls]} — 가능한 만큼만")
        sel = select_for_class(pool, need[cls], args.seed) if need[cls] else []
        new_sel[cls] = sel
        print(f"  {cls}: 신규후보 {len(pool)} → 선택 {len(sel)}")

    # 신규 작품 썸네일 다운로드 (공개·소용량)
    print(f"\n신규 썸네일 다운로드 (workers={args.workers})...")
    dl = download_all(new_sel, out_dir, args.workers, args.seed)

    # 결합 매니페스트 작성 (컬럼 순서 = 기존 헤더 유지)
    manifest_meta_cols = [c for c in METADATA_COLUMNS if c not in MANIFEST_EXCLUDE]
    columns = manifest_meta_cols + EXTRA_COLUMNS
    assert columns == hdr, f"컬럼 불일치:\n cols={columns}\n hdr={hdr}"
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "manifest"
    ws.append(columns)
    for r in existing:                       # 기존 보존
        ws.append([r.get(c, "") for c in columns])
    n_new = 0
    for cls in target:                       # 신규 추가
        for rec in new_sel[cls]:
            idx = rec["원문인덱스"]; d = dl.get(idx, {})
            row = [rec.get(k, "") for k in manifest_meta_cols]
            row += [d.get("local_path", ""), d.get("download_status", ""),
                    d.get("media_available", "False"), SUBFOLDER.get(cls, "misc")]
            ws.append(row); n_new += 1

    backup = out_dir / "manifest_1000_backup.xlsx"
    if not backup.exists():
        shutil.copy(manifest_path, backup); print(f"기존 매니페스트 백업: {backup}")
    wb.save(manifest_path)
    print(f"\n새 manifest.xlsx: 기존 {len(existing)} + 신규 {n_new} = {len(existing)+n_new}행")
    # 최종 분류 분포
    final = Counter(r["분류"] for r in existing)
    for cls in target:
        final[cls] += len(new_sel[cls])
    print(f"최종 분포: {dict(final)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

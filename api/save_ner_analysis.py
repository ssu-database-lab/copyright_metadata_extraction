#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NER 라벨 추출 현황을 CSV/Excel로 저장
"""

import csv
from pathlib import Path

def save_ner_labels_to_csv():
    """NER 라벨 현황을 CSV로 저장"""
    
    # 23개 라벨 (정의된 ENTITY_TYPES)
    ALL_ENTITY_TYPES = [
        "NAME", "PHONE", "ADDRESS", "DATE", "COMPANY", "EMAIL", "POSITION",
        "CONTRACT_TYPE", "CONSENT_TYPE", "RIGHT_INFO", "MONEY", "PERIOD",
        "PROJECT_NAME", "LAW_REFERENCE", "ID_NUM", "TITLE", "URL",
        "DESCRIPTION", "TYPE", "STATUS", "DEPARTMENT", "LANGUAGE", "QUANTITY"
    ]
    
    # 훈련 데이터 경로
    training_dir = Path("/home/peppermint/copyright_metadata_extraction/api/src/ner/training/google-bert-bert-base-multilingual-cased")
    
    def extract_labels_from_file(file_path):
        """BIO 형식 파일에서 라벨 추출"""
        labels_found = set()
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line and '\t' in line:
                        parts = line.split('\t')
                        if len(parts) == 2:
                            token, label = parts
                            if label != 'O':
                                entity_type = label.split('-')[1]
                                labels_found.add(entity_type)
        except Exception as e:
            print(f"❌ 파일 읽기 오류: {file_path} - {e}")
        
        return labels_found
    
    # 각 파일에서 라벨 추출
    train_labels = extract_labels_from_file(training_dir / "train.txt")
    val_labels = extract_labels_from_file(training_dir / "validation.txt")
    test_labels = extract_labels_from_file(training_dir / "test.txt")
    
    # 전체 사용된 라벨
    all_used_labels = train_labels | val_labels | test_labels
    
    # CSV 데이터 준비
    csv_data = []
    csv_data.append(["#", "라벨명", "Train", "Validation", "Test", "최종"])  # 헤더
    
    extracted_count = 0
    for idx, entity_type in enumerate(ALL_ENTITY_TYPES, 1):
        in_train = "O" if entity_type in train_labels else "X"
        in_val = "O" if entity_type in val_labels else "X"
        in_test = "O" if entity_type in test_labels else "X"
        in_all = "O" if entity_type in all_used_labels else "X"
        
        if in_all == "O":
            extracted_count += 1
        
        csv_data.append([
            str(idx),
            entity_type,
            in_train,
            in_val,
            in_test,
            in_all
        ])
    
    # 요약 정보 추가
    csv_data.append([])  # 빈 줄
    csv_data.append(["최종 결과"])
    csv_data.append(["항목", "개수", "비율"])
    csv_data.append(["총 정의된 라벨", str(len(ALL_ENTITY_TYPES)), "100.0%"])
    csv_data.append(["훈련 데이터에 포함된 라벨", str(extracted_count), f"{(extracted_count/len(ALL_ENTITY_TYPES)*100):.1f}%"])
    csv_data.append(["미포함 라벨", str(len(ALL_ENTITY_TYPES) - extracted_count), f"{((len(ALL_ENTITY_TYPES) - extracted_count)/len(ALL_ENTITY_TYPES)*100):.1f}%"])
    
    csv_data.append([])  # 빈 줄
    csv_data.append(["추출된 라벨 (16개)"])
    extracted_labels_list = sorted([et for et in ALL_ENTITY_TYPES if et in all_used_labels])
    csv_data.append([", ".join(extracted_labels_list)])
    
    csv_data.append([])  # 빈 줄
    csv_data.append(["미추출 라벨 (7개)"])
    not_extracted = sorted([et for et in ALL_ENTITY_TYPES if et not in all_used_labels])
    csv_data.append([", ".join(not_extracted)])
    
    # 홈 디렉토리에 저장
    home_dir = Path.home()
    csv_file = home_dir / "ner_labels_analysis.csv"
    
    try:
        with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerows(csv_data)
        
        print(f"✅ CSV 파일 저장 완료!")
        print(f"📁 경로: {csv_file}")
        print(f"📊 내용: NER 라벨 추출 현황 분석")
        return True
    except Exception as e:
        print(f"❌ CSV 저장 실패: {e}")
        return False


def try_save_excel():
    """Excel 저장 시도"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 23개 라벨
        ALL_ENTITY_TYPES = [
            "NAME", "PHONE", "ADDRESS", "DATE", "COMPANY", "EMAIL", "POSITION",
            "CONTRACT_TYPE", "CONSENT_TYPE", "RIGHT_INFO", "MONEY", "PERIOD",
            "PROJECT_NAME", "LAW_REFERENCE", "ID_NUM", "TITLE", "URL",
            "DESCRIPTION", "TYPE", "STATUS", "DEPARTMENT", "LANGUAGE", "QUANTITY"
        ]
        
        training_dir = Path("/home/peppermint/copyright_metadata_extraction/api/src/ner/training/google-bert-bert-base-multilingual-cased")
        
        def extract_labels_from_file(file_path):
            labels_found = set()
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line and '\t' in line:
                            parts = line.split('\t')
                            if len(parts) == 2:
                                token, label = parts
                                if label != 'O':
                                    entity_type = label.split('-')[1]
                                    labels_found.add(entity_type)
            except:
                pass
            return labels_found
        
        train_labels = extract_labels_from_file(training_dir / "train.txt")
        val_labels = extract_labels_from_file(training_dir / "validation.txt")
        test_labels = extract_labels_from_file(training_dir / "test.txt")
        all_used_labels = train_labels | val_labels | test_labels
        
        # Excel 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "NER Labels"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 작성
        headers = ["#", "라벨명", "Train", "Validation", "Test", "최종"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # 데이터 작성
        extracted_count = 0
        for idx, entity_type in enumerate(ALL_ENTITY_TYPES, 1):
            in_train = "O" if entity_type in train_labels else "X"
            in_val = "O" if entity_type in val_labels else "X"
            in_test = "O" if entity_type in test_labels else "X"
            in_all = "O" if entity_type in all_used_labels else "X"
            
            if in_all == "O":
                extracted_count += 1
            
            row = idx + 1
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = entity_type
            ws.cell(row=row, column=3).value = in_train
            ws.cell(row=row, column=4).value = in_val
            ws.cell(row=row, column=5).value = in_test
            ws.cell(row=row, column=6).value = in_all
            
            # 스타일 적용 (O는 초록색, X는 빨간색)
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                if col >= 3:  # Train, Validation, Test, 최종
                    cell_value = cell.value
                    if cell_value == "O":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif cell_value == "X":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        
        # 홈 디렉토리에 저장
        home_dir = Path.home()
        excel_file = home_dir / "ner_labels_analysis.xlsx"
        
        wb.save(excel_file)
        print(f"✅ Excel 파일 저장 완료!")
        print(f"📁 경로: {excel_file}")
        print(f"📊 내용: NER 라벨 추출 현황 분석 (색상 코딩: 초록=O, 빨강=X)")
        return True
        
    except ImportError:
        return False


if __name__ == "__main__":
    print("=" * 80)
    print("📊 NER 라벨 분석 데이터 저장")
    print("=" * 80)
    
    # Excel 시도
    if try_save_excel():
        pass
    else:
        print("⚠️  openpyxl 라이브러리가 없어 CSV로 저장합니다...")
        save_ner_labels_to_csv()
